"""Phase 2-A/B: 年级 + 个人学科报告导出端点测试。"""
import pytest

from edu_cloud.models.school import School
from edu_cloud.models.user import User
from edu_cloud.models.user_role import UserRole
from edu_cloud.models.exam import Exam, Subject, Question
from edu_cloud.modules.scan.models import StudentAnswer
from edu_cloud.modules.grading.models import GradingTask, GradingResult
from edu_cloud.modules.student.models import Class, Student
from edu_cloud.shared.auth import create_access_token


@pytest.fixture
async def export_setup(client, db):
    school = School(name="ES", code="ES01")
    db.add(school)
    await db.commit()
    user = User(username="exporter", display_name="导出测试")
    user.set_password("p")
    db.add(user)
    await db.commit()
    db.add(UserRole(user_id=user.id, role="admin", school_id=school.id, is_primary=True))
    await db.flush()
    token = create_access_token({"sub": user.id, "school_id": school.id, "role": "admin"})
    headers = {"Authorization": f"Bearer {token}"}

    exam = Exam(name="期中考试", school_id=school.id)
    db.add(exam)
    await db.commit()
    subject = Subject(exam_id=exam.id, name="数学", code="math", school_id=school.id)
    db.add(subject)
    await db.commit()

    q1 = Question(subject_id=subject.id, name="1", question_type="choice",
                  max_score=5.0, school_id=school.id)
    q2 = Question(subject_id=subject.id, name="2", question_type="essay",
                  max_score=15.0, school_id=school.id)
    db.add_all([q1, q2])
    await db.commit()

    task = GradingTask(
        subject_id=subject.id, school_id=school.id,
        status="completed", total=10, completed=10, failed=0, created_by=user.id,
    )
    db.add(task)
    await db.commit()

    klass = Class(name="一班", grade="高二", school_id=school.id)
    db.add(klass)
    await db.commit()

    students = []
    student_scores = [
        ("张三", "S001", [5, 14]),
        ("李四", "S002", [4, 12]),
        ("王五", "S003", [3, 10]),
        ("赵六", "S004", [5, 8]),
        ("钱七", "S005", [2, 6]),
    ]
    for name, num, scores in student_scores:
        student = Student(
            name=name, student_number=num, class_id=klass.id, school_id=school.id,
        )
        db.add(student)
        await db.commit()
        students.append(student)
        for q, score in zip([q1, q2], scores):
            a = StudentAnswer(
                exam_id=exam.id, subject_id=subject.id, student_id=student.id,
                question_id=q.id, image_path=f"/fake/{student.id}_{q.id}.png", school_id=school.id,
            )
            db.add(a)
            await db.commit()
            r = GradingResult(
                ai_task_id=task.id, answer_id=a.id, question_id=q.id,
                school_id=school.id, ai_score=float(score), final_score=float(score),
                max_score=q.max_score, ai_feedback="f", ai_confidence=0.9, status="ai_done",
            )
            db.add(r)
            await db.commit()

    return {
        "headers": headers, "exam_id": exam.id, "subject_id": subject.id,
        "school_id": school.id, "class_id": klass.id,
        "students": students,
    }


async def test_export_pdf_returns_200_with_pdf_content_type(client, export_setup):
    resp = await client.get(
        f"/api/v1/analytics/report/grade/{export_setup['exam_id']}/{export_setup['subject_id']}/export"
        "?format=pdf",
        headers=export_setup["headers"],
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/pdf"
    body = resp.content
    assert body.startswith(b"%PDF"), "应为 PDF 文件头"
    assert len(body) > 1000, "PDF 文件不应过小"
    assert "Content-Disposition" in resp.headers
    assert "attachment" in resp.headers["Content-Disposition"]


async def test_export_xlsx_returns_200_with_xlsx_content_type(client, export_setup):
    resp = await client.get(
        f"/api/v1/analytics/report/grade/{export_setup['exam_id']}/{export_setup['subject_id']}/export"
        "?format=xlsx",
        headers=export_setup["headers"],
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    body = resp.content
    assert body[:2] == b"PK", "XLSX 应为 ZIP 文件头"
    assert len(body) > 1000


async def test_export_xlsx_has_expected_sheets(client, export_setup):
    """openpyxl 解出的 sheet 列表应覆盖 4 个维度。"""
    from openpyxl import load_workbook
    from io import BytesIO

    resp = await client.get(
        f"/api/v1/analytics/report/grade/{export_setup['exam_id']}/{export_setup['subject_id']}/export"
        "?format=xlsx",
        headers=export_setup["headers"],
    )
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content))
    assert set(wb.sheetnames) == {"总览", "分数段分布", "班级对比", "题目分析"}
    # 总览 sheet 含考试名 + 科目名
    overview = wb["总览"]
    cells = [overview.cell(row=r, column=2).value for r in range(1, 9)]
    assert "期中考试" in cells
    assert "数学" in cells


async def test_export_invalid_format_400(client, export_setup):
    resp = await client.get(
        f"/api/v1/analytics/report/grade/{export_setup['exam_id']}/{export_setup['subject_id']}/export"
        "?format=docx",
        headers=export_setup["headers"],
    )
    assert resp.status_code == 422  # Pydantic Query pattern 校验


async def test_export_404_when_exam_missing(client, export_setup):
    resp = await client.get(
        f"/api/v1/analytics/report/grade/no-such-exam/{export_setup['subject_id']}/export"
        "?format=pdf",
        headers=export_setup["headers"],
    )
    assert resp.status_code == 404


async def test_export_404_when_subject_not_in_exam(client, export_setup):
    resp = await client.get(
        f"/api/v1/analytics/report/grade/{export_setup['exam_id']}/no-such-subject/export"
        "?format=pdf",
        headers=export_setup["headers"],
    )
    assert resp.status_code == 404


# ── Phase 2-B: 个人学科报告 ────────────────────────────────────


async def test_export_student_pdf_returns_200(client, export_setup):
    student = export_setup["students"][0]
    resp = await client.get(
        f"/api/v1/analytics/report/student/{student.id}/{export_setup['exam_id']}/"
        f"{export_setup['subject_id']}/export?format=pdf",
        headers=export_setup["headers"],
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content.startswith(b"%PDF")
    assert len(resp.content) > 1000


async def test_export_student_xlsx_has_three_sheets(client, export_setup):
    from openpyxl import load_workbook
    from io import BytesIO

    student = export_setup["students"][0]
    resp = await client.get(
        f"/api/v1/analytics/report/student/{student.id}/{export_setup['exam_id']}/"
        f"{export_setup['subject_id']}/export?format=xlsx",
        headers=export_setup["headers"],
    )
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content))
    assert set(wb.sheetnames) == {"个人成绩", "各题得失分", "薄弱题"}
    overview = wb["个人成绩"]
    cells = [overview.cell(row=r, column=2).value for r in range(1, 11)]
    assert "张三" in cells
    assert "S001" in cells
    assert "数学" in cells


async def test_export_student_404_unknown_student(client, export_setup):
    resp = await client.get(
        f"/api/v1/analytics/report/student/no-such-student/{export_setup['exam_id']}/"
        f"{export_setup['subject_id']}/export?format=pdf",
        headers=export_setup["headers"],
    )
    assert resp.status_code == 404


async def test_export_student_403_when_class_outside_scope(client, db, export_setup):
    """班主任只能看自己班级的学生。"""
    user2 = User(username="other_homeroom", display_name="另一班班主任")
    user2.set_password("p")
    db.add(user2)
    await db.commit()
    other_class = Class(name="二班", grade="高二", school_id=export_setup["school_id"])
    db.add(other_class)
    await db.commit()
    role = UserRole(
        user_id=user2.id, role="homeroom_teacher",
        school_id=export_setup["school_id"], is_primary=True,
        class_ids=[other_class.id],
    )
    db.add(role)
    await db.commit()
    token = create_access_token({
        "sub": user2.id, "school_id": export_setup["school_id"], "role": "homeroom_teacher",
    })
    headers = {"Authorization": f"Bearer {token}"}

    student = export_setup["students"][0]  # 一班学生
    resp = await client.get(
        f"/api/v1/analytics/report/student/{student.id}/{export_setup['exam_id']}/"
        f"{export_setup['subject_id']}/export?format=pdf",
        headers=headers,
    )
    assert resp.status_code in (403, 404), resp.text


async def test_export_403_when_subject_outside_visible_scope(client, db, export_setup):
    """非授权角色（subject_teacher 仅教语文）请求数学应 403。"""
    user2 = User(username="yw_only", display_name="语文老师")
    user2.set_password("p")
    db.add(user2)
    await db.commit()
    role = UserRole(
        user_id=user2.id, role="subject_teacher",
        school_id=export_setup["school_id"], is_primary=True,
        subject_codes=["chinese"],  # 仅教语文
    )
    db.add(role)
    await db.commit()
    token = create_access_token({
        "sub": user2.id, "school_id": export_setup["school_id"], "role": "subject_teacher",
    })
    headers = {"Authorization": f"Bearer {token}"}

    resp = await client.get(
        f"/api/v1/analytics/report/grade/{export_setup['exam_id']}/{export_setup['subject_id']}/export"
        "?format=pdf",
        headers=headers,
    )
    assert resp.status_code in (403, 404), resp.text
