"""Admin config API tests for conduct module."""
import pytest

from edu_cloud.modules.conduct.models import ConductClassConfig, StudentProfile
from edu_cloud.modules.conduct.crypto import encrypt
from edu_cloud.models.guardian import GuardianStudentLink
from edu_cloud.models.user import User
from edu_cloud.models.user_role import UserRole


@pytest.mark.anyio
async def test_get_config(client, db, conduct_config, school_class_student, homeroom_headers):
    """Homeroom teacher can get conduct config for their class."""
    _, cls, _ = school_class_student
    resp = await client.get(
        f"/api/v1/conduct/classes/{cls.id}/config",
        headers=homeroom_headers,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["class_id"] == cls.id
    assert data["invite_code"] == "TEST01"
    assert data["verify_code_type"] == "custom"
    assert data["is_active"] is True


@pytest.mark.anyio
async def test_update_config(client, db, conduct_config, school_class_student, homeroom_headers):
    """Homeroom teacher can update verify_code_type."""
    _, cls, _ = school_class_student
    resp = await client.put(
        f"/api/v1/conduct/classes/{cls.id}/config",
        headers=homeroom_headers,
        json={"verify_code_type": "phone"},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["verify_code_type"] == "phone"
    assert data["invite_code"] == "TEST01"  # unchanged


@pytest.mark.anyio
async def test_regenerate_invite_code(client, db, conduct_config, school_class_student, homeroom_headers):
    """Homeroom teacher gets a new invite code (different from old one)."""
    _, cls, _ = school_class_student
    resp = await client.post(
        f"/api/v1/conduct/classes/{cls.id}/config/regenerate-code",
        headers=homeroom_headers,
    )
    assert resp.status_code == 200
    new_code = resp.json()["invite_code"]
    assert new_code  # non-empty
    assert new_code != "TEST01"  # different from original


@pytest.mark.anyio
async def test_list_parents(client, db, conduct_config, school_class_student, homeroom_headers):
    """After a parent registers and binds, list shows them."""
    school, cls, student = school_class_student

    # Set up student verify code for binding
    profile = StudentProfile(student_id=student.id, verify_code=encrypt("ADMIN01"))
    db.add(profile)
    await db.commit()

    # Register parent via API
    resp = await client.post("/api/v1/conduct/parent/register", json={
        "invite_code": "TEST01",
        "display_name": "管理测试家长",
        "phone": "13900000001",
        "password": "abc123",
    })
    assert resp.status_code == 200
    token = resp.json()["access_token"]
    parent_headers = {"Authorization": f"Bearer {token}"}

    # Bind parent to student
    resp = await client.post("/api/v1/conduct/parent/bind", headers=parent_headers, json={
        "class_id": cls.id,
        "student_name": "张三",
        "verify_code": "ADMIN01",
        "relationship": "father",
    })
    assert resp.status_code == 200

    # Admin lists parents
    resp = await client.get(
        f"/api/v1/conduct/classes/{cls.id}/parents",
        headers=homeroom_headers,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) == 1
    assert data[0]["display_name"] == "管理测试家长"
    assert data[0]["phone"] == "13900000001"
    assert len(data[0]["students"]) == 1
    assert data[0]["students"][0]["student_name"] == "张三"
    assert data[0]["students"][0]["relationship"] == "father"


@pytest.mark.anyio
async def test_remove_parent(client, db, conduct_config, school_class_student, homeroom_headers):
    """Remove parent, then list no longer shows them."""
    school, cls, student = school_class_student

    # Set up student verify code for binding
    profile = StudentProfile(student_id=student.id, verify_code=encrypt("ADMIN02"))
    db.add(profile)
    await db.commit()

    # Register parent via API
    resp = await client.post("/api/v1/conduct/parent/register", json={
        "invite_code": "TEST01",
        "display_name": "待删家长",
        "phone": "13900000002",
        "password": "abc123",
    })
    assert resp.status_code == 200
    token = resp.json()["access_token"]
    parent_headers = {"Authorization": f"Bearer {token}"}

    # Bind parent to student
    resp = await client.post("/api/v1/conduct/parent/bind", headers=parent_headers, json={
        "class_id": cls.id,
        "student_name": "张三",
        "verify_code": "ADMIN02",
        "relationship": "mother",
    })
    assert resp.status_code == 200

    # List parents to get user_id
    resp = await client.get(
        f"/api/v1/conduct/classes/{cls.id}/parents",
        headers=homeroom_headers,
    )
    assert resp.status_code == 200
    parents = resp.json()
    assert len(parents) == 1
    parent_user_id = parents[0]["user_id"]

    # Remove parent
    resp = await client.delete(
        f"/api/v1/conduct/classes/{cls.id}/parents/{parent_user_id}",
        headers=homeroom_headers,
    )
    assert resp.status_code == 200
    assert resp.json()["deleted"] == 1

    # List parents again — should be empty
    resp = await client.get(
        f"/api/v1/conduct/classes/{cls.id}/parents",
        headers=homeroom_headers,
    )
    assert resp.status_code == 200
    assert resp.json() == []


# ── F006 Round 3: Excel 导出入口级测试（openpyxl 解包内容断言） ──

@pytest.mark.anyio
async def test_export_records_excel(client, db, homeroom_headers, school_class_student, homeroom_teacher):
    """F006 R3: 导出 Excel 解包后必须含 header + N 条记录行 + 字段正确（不被 inner join 过滤）."""
    from datetime import date as _date
    from io import BytesIO
    from openpyxl import load_workbook
    from edu_cloud.modules.conduct.models import ConductRecord

    school, cls, student = school_class_student
    # 关键修复：operator_id 必须对应真实 User（inner join 不过滤）
    operator_id = homeroom_teacher.id

    rec1 = ConductRecord(
        student_id=student.id, class_id=cls.id, points=3,
        reason="课堂表现好", date=_date(2026, 4, 10),
        operator_id=operator_id, source="manual",
    )
    rec2 = ConductRecord(
        student_id=student.id, class_id=cls.id, points=-2,
        reason="迟到", date=_date(2026, 4, 11),
        operator_id=operator_id, source="manual",
    )
    db.add_all([rec1, rec2])
    await db.commit()

    resp = await client.get(
        f"/api/v1/conduct/classes/{cls.id}/export/records",
        headers=homeroom_headers,
    )
    assert resp.status_code == 200, f"导出失败: {resp.text[:200]}"
    assert resp.content[:4] == b"PK\x03\x04", "非 xlsx 魔数"

    # openpyxl 解包断言行数 + 字段内容
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 3, f"应有 header + 2 记录行，实际 {len(rows)}"
    assert rows[0] == ("日期", "学生姓名", "积分", "原因", "操作人", "来源")

    # 验证记录行内容（order_by date desc → rec2 先，rec1 后）
    body_rows = rows[1:]
    reasons = {r[3] for r in body_rows}
    assert reasons == {"课堂表现好", "迟到"}, f"原因字段不匹配: {reasons}"
    points_values = {r[2] for r in body_rows}
    assert points_values == {3, -2}, f"积分字段不匹配: {points_values}"
    student_names = {r[1] for r in body_rows}
    assert student_names == {"张三"}
    operators = {r[4] for r in body_rows}
    assert operators == {"teacher_hr"}, f"操作人字段不匹配（inner join 过滤警告）: {operators}"


@pytest.mark.anyio
async def test_export_rankings_excel(client, db, homeroom_headers, school_class_student, homeroom_teacher):
    """F007 (R3 GPT-reported test-gap): 排行榜必须覆盖"排序 + 聚合"两个核心语义。

    单学生、零积分的旧断言无法在错误实现（比如破坏 SUM 或 ORDER BY DESC）下失败。
    本测试构造 2 名学生 + 多条不同积分记录，断言：
      1. 学生 B (积分 +7+3=10) 排第 1
      2. 学生 A (积分 +2-5=-3) 排第 2
      3. 聚合值正确（SUM points per student）
    """
    from datetime import date as _date
    from io import BytesIO
    from openpyxl import load_workbook
    from edu_cloud.modules.student.models import Student
    from edu_cloud.modules.conduct.models import ConductRecord

    school, cls, student_a = school_class_student
    # 第二名学生（同班）
    student_b = Student(
        name="李四", student_number="2026002",
        class_id=cls.id, school_id=school.id,
    )
    db.add(student_b)
    await db.flush()

    operator_id = homeroom_teacher.id
    # 学生 A: 2 - 5 = -3
    # 学生 B: 7 + 3 = 10（总分更高，应排第 1）
    records = [
        ConductRecord(student_id=student_a.id, class_id=cls.id, points=2,
                      reason="A +2", date=_date(2026, 4, 10),
                      operator_id=operator_id, source="manual"),
        ConductRecord(student_id=student_a.id, class_id=cls.id, points=-5,
                      reason="A -5", date=_date(2026, 4, 11),
                      operator_id=operator_id, source="manual"),
        ConductRecord(student_id=student_b.id, class_id=cls.id, points=7,
                      reason="B +7", date=_date(2026, 4, 10),
                      operator_id=operator_id, source="manual"),
        ConductRecord(student_id=student_b.id, class_id=cls.id, points=3,
                      reason="B +3", date=_date(2026, 4, 11),
                      operator_id=operator_id, source="manual"),
    ]
    db.add_all(records)
    await db.commit()

    resp = await client.get(
        f"/api/v1/conduct/classes/{cls.id}/export/rankings",
        headers=homeroom_headers,
    )
    assert resp.status_code == 200, f"排行导出失败: {resp.text[:200]}"
    assert resp.content[:4] == b"PK\x03\x04", "非 xlsx 魔数"

    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # header + 2 学生
    assert len(rows) == 3, f"应有 header + 2 学生行，实际 {len(rows)}"
    assert rows[0] == ("排名", "学生姓名", "学号", "总积分")

    # 排序断言：第 1 行是学生 B（总分 10），第 2 行是学生 A（总分 -3）
    assert rows[1][0] == 1, f"排名 1 错误: {rows[1]}"
    assert rows[1][1] == "李四", f"第 1 名应为学生 B(李四)，实际: {rows[1]}"
    assert rows[1][2] == "2026002"
    assert rows[1][3] == 10, f"学生 B 聚合总分应为 10(=7+3)，实际 {rows[1][3]}"

    assert rows[2][0] == 2, f"排名 2 错误: {rows[2]}"
    assert rows[2][1] == "张三", f"第 2 名应为学生 A(张三)，实际: {rows[2]}"
    assert rows[2][2] == "2026001"
    assert rows[2][3] == -3, f"学生 A 聚合总分应为 -3(=2-5)，实际 {rows[2][3]}"

    # 显式断言 ORDER BY desc：总分 B > 总分 A
    assert rows[1][3] > rows[2][3], "排序错误：排名 1 总分应大于排名 2"


# ── F002: Class-scope 守卫红测 ──

@pytest.mark.anyio
async def test_class_scope_denies_other_class(client, db, homeroom_headers, school_class_student):
    """F002: homeroom_teacher of class A 访问 class B 的 config → 403."""
    school, cls_a, _ = school_class_student
    # 创建另一个班 B（同校，但 homeroom_teacher 不在其 class_ids 中）
    from edu_cloud.modules.student.models import Class
    cls_b = Class(name="高一(2)班", grade="高一", grade_number=1, school_id=school.id)
    db.add(cls_b)
    await db.commit()

    resp = await client.get(
        f"/api/v1/conduct/classes/{cls_b.id}/config",
        headers=homeroom_headers,
    )
    assert resp.status_code == 403, f"跨班访问未被拒绝: {resp.status_code} {resp.text[:100]}"
    assert "scope" in resp.text.lower() or "class" in resp.text.lower()


@pytest.mark.anyio
async def test_resource_affinity_denies_cross_class_rule(client, db, homeroom_headers, school_class_student):
    """F002: 用外班的 rule_category_id 在本班路径下更新 → 404（资源不属于本班）."""
    school, cls_a, _ = school_class_student
    # 为班 A 的 homeroom_teacher 配置 class_ids (fixture 已做)
    # 创建班 B 和它的规则分类
    from edu_cloud.modules.student.models import Class
    from edu_cloud.modules.conduct.models import ConductRuleCategory
    cls_b = Class(name="高一(2)班", grade="高一", grade_number=1, school_id=school.id)
    db.add(cls_b)
    await db.flush()
    cat_b = ConductRuleCategory(name="B班规则", class_id=cls_b.id, scope="class")
    db.add(cat_b)
    await db.commit()

    # 用 cls_a.id (本班) + cat_b.id (外班资源) 调用 PUT
    resp = await client.put(
        f"/api/v1/conduct/classes/{cls_a.id}/rules/categories/{cat_b.id}",
        headers=homeroom_headers,
        json={"name": "恶意修改", "sort_order": 99},
    )
    # 预期：要么 403（scope 拦截早于资源查找），要么 404（资源不属于本班）
    assert resp.status_code in (403, 404), f"跨班资源越权未被拒绝: {resp.status_code}"


@pytest.mark.anyio
async def test_resource_affinity_denies_cross_class_group(client, db, homeroom_headers, school_class_student):
    """F002: 用外班的 group_id 在本班路径下删除 → 404."""
    school, cls_a, _ = school_class_student
    from edu_cloud.modules.student.models import Class
    from edu_cloud.modules.conduct.models import ConductGroup
    cls_b = Class(name="高一(3)班", grade="高一", grade_number=1, school_id=school.id)
    db.add(cls_b)
    await db.flush()
    group_b = ConductGroup(name="B 组", class_id=cls_b.id)
    db.add(group_b)
    await db.commit()

    resp = await client.delete(
        f"/api/v1/conduct/classes/{cls_a.id}/groups/{group_b.id}",
        headers=homeroom_headers,
    )
    assert resp.status_code in (403, 404), f"跨班小组越权未被拒绝: {resp.status_code}"


# ── F002 Round 3: body-field / batch-write 越权红测 ──

@pytest.mark.anyio
async def test_add_points_cross_class_rule_item_rejected(
    client, db, homeroom_headers, school_class_student,
):
    """F002 R3: 本班路径 + 外班 rule_item_id → 404（check_rule_item_class 拦截）."""
    from edu_cloud.modules.student.models import Class
    from edu_cloud.modules.conduct.models import ConductRuleCategory, ConductRuleItem

    school, cls_a, student_a = school_class_student
    # 班 B 及其规则条目
    cls_b = Class(name="高一(2)班", grade="高一", grade_number=1, school_id=school.id)
    db.add(cls_b)
    await db.flush()
    cat_b = ConductRuleCategory(name="B班规则", class_id=cls_b.id, scope="class")
    db.add(cat_b)
    await db.flush()
    item_b = ConductRuleItem(name="B班条目", points=-2, category_id=cat_b.id)
    db.add(item_b)
    await db.commit()

    # homeroom_teacher(班A) 用班 A 路径 + 外班 rule_item_id 加分
    resp = await client.post(
        f"/api/v1/conduct/classes/{cls_a.id}/records",
        headers=homeroom_headers,
        json={
            "student_ids": [student_a.id],
            "points": -2,
            "reason": "跨班 rule_item_id 注入",
            "rule_item_id": item_b.id,
        },
    )
    assert resp.status_code == 404, (
        f"跨班 rule_item_id 未被拒绝: status={resp.status_code} body={resp.text[:200]}"
    )


@pytest.mark.anyio
async def test_add_group_members_cross_class_student_rejected(
    client, db, homeroom_headers, school_class_student,
):
    """F002 R3: 本班 group_id + 外班 student_ids → 404（check_students_class 拦截）."""
    from edu_cloud.modules.student.models import Class, Student
    from edu_cloud.modules.conduct.models import ConductGroup

    school, cls_a, _ = school_class_student
    # 班 A 的 group
    group_a = ConductGroup(name="A 组", class_id=cls_a.id)
    db.add(group_a)
    await db.flush()
    # 班 B 及其学生
    cls_b = Class(name="高一(2)班", grade="高一", grade_number=1, school_id=school.id)
    db.add(cls_b)
    await db.flush()
    student_b = Student(name="李四", student_number="2026099", class_id=cls_b.id, school_id=school.id)
    db.add(student_b)
    await db.commit()

    resp = await client.post(
        f"/api/v1/conduct/classes/{cls_a.id}/groups/{group_a.id}/members",
        headers=homeroom_headers,
        json={"student_ids": [student_b.id]},
    )
    assert resp.status_code == 404, (
        f"跨班 student_ids 未被拒绝: status={resp.status_code} body={resp.text[:200]}"
    )


@pytest.mark.anyio
async def test_remove_group_member_cross_class_student_rejected(
    client, db, homeroom_headers, school_class_student,
):
    """F002 R3: 本班 group_id + 外班 student_id path → 404（check_students_class 拦截）."""
    from edu_cloud.modules.student.models import Class, Student
    from edu_cloud.modules.conduct.models import ConductGroup

    school, cls_a, _ = school_class_student
    group_a = ConductGroup(name="A 组", class_id=cls_a.id)
    db.add(group_a)
    await db.flush()
    cls_b = Class(name="高一(3)班", grade="高一", grade_number=1, school_id=school.id)
    db.add(cls_b)
    await db.flush()
    student_b = Student(name="王五", student_number="2026098", class_id=cls_b.id, school_id=school.id)
    db.add(student_b)
    await db.commit()

    resp = await client.delete(
        f"/api/v1/conduct/classes/{cls_a.id}/groups/{group_a.id}/members/{student_b.id}",
        headers=homeroom_headers,
    )
    assert resp.status_code == 404, (
        f"跨班 student_id (path) 未被拒绝: status={resp.status_code} body={resp.text[:200]}"
    )
