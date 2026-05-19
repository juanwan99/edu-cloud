"""Excel parser for external exam data (joint exams like 天一大联考).

Three entry points:
- parse_question_scores_xlsx  -- per-question horizontal table
- parse_totals_xlsx           -- subject totals horizontal table
- parse_zip                   -- zip archive containing xlsx files
"""

from __future__ import annotations

import logging
import os
import re
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

# ── subject mappings ──────────────────────────────────────────────

SUBJECT_CODE_MAP: dict[str, str] = {
    "语文": "YW",
    "数学": "SX",
    "英语": "YY",
    "物理": "WL",
    "化学": "HX",
    "生物": "SW",
    "历史": "LS",
    "地理": "DL",
    "政治": "ZZ",
}

SUBJECT_MAX_SCORE: dict[str, float] = {
    "YW": 150,
    "SX": 150,
    "YY": 150,
    "WL": 100,
    "HX": 100,
    "SW": 100,
    "LS": 100,
    "DL": 100,
    "ZZ": 100,
}

# ── data classes ──────────────────────────────────────────────────


@dataclass
class QuestionDef:
    name: str
    question_type: str  # choice / multi_choice / essay / unknown / synthetic
    max_score: float
    correct_answer: str | None = None
    max_score_inferred: bool = False


@dataclass
class StudentScore:
    student_key: str
    student_name: str
    class_name: str | None = None
    school_name: str | None = None
    elective: str | None = None
    raw_total: float | None = None
    converted_score: float | None = None  # 赋分
    grade_level: str | None = None
    class_rank: int | None = None
    school_rank: int | None = None
    objective_subtotal: float | None = None
    subjective_subtotal: float | None = None
    question_scores: dict[str, float] = field(default_factory=dict)
    is_absent: bool = False


@dataclass
class ParsedSubjectData:
    subject_name: str
    subject_code: str
    questions: list[QuestionDef] = field(default_factory=list)
    students: list[StudentScore] = field(default_factory=list)


@dataclass
class ParsedExamData:
    subjects: list[ParsedSubjectData] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


# ── helpers ───────────────────────────────────────────────────────

_FIXED_COLS = {"学校", "班级", "考号", "姓名", "选科"}
_AGGREGATE_HEADERS = {"赋分", "原始分", "等级", "班名次", "校名次", "客观题", "主观题"}

# Headers that identify fixed (non-subject) columns in totals sheets.
_TOTALS_FIXED = {
    "考号", "学号", "准考证号", "姓名", "班级", "选科",
    "总分", "班名次", "校名次", "年级名次", "学校",
}


def _to_float(val: object) -> float | None:
    """Coerce a cell value to float.  Returns None on failure."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.strip()
        if val == "" or val == "-":
            return None
        try:
            return float(val)
        except ValueError:
            return None
    return None


def _to_int(val: object) -> int | None:
    f = _to_float(val)
    if f is None:
        return None
    return int(f)


def _detect_subject_from_filename(filename: str) -> tuple[str, str] | None:
    """Try to extract a subject name from the file path/name."""
    stem = Path(filename).stem
    for cn_name, code in SUBJECT_CODE_MAP.items():
        if cn_name in stem:
            return cn_name, code
    return None


def _infer_question_type(name: str) -> str:
    if re.match(r"选择\d+", name):
        return "choice"
    return "essay"


# ── parse_question_scores_xlsx ────────────────────────────────────


def parse_question_scores_xlsx(file_path: str | Path) -> ParsedExamData:
    """Parse a per-question-score xlsx (horizontal table).

    Expected layout (real data from 天一大联考):
      Row 1: 学校 | 班级 | 考号 | 姓名 | 选科 | <subject>(merged) | 客观题(merged) | 主观题(merged)
      Row 2: (blanks for first 5) | 赋分 | 原始分 | 等级 | 班名次 | 校名次 | 客观题 | 主观题 | 选择1 | 选择2 | ... | 17 | 18 | ...
      Row 3+: data
    """
    file_path = Path(file_path)
    wb = openpyxl.load_workbook(str(file_path))
    ws = wb.active
    warnings: list[str] = []

    rows = list(ws.iter_rows(values_only=False))
    if len(rows) < 3:
        return ParsedExamData(warnings=["Sheet has fewer than 3 rows"])

    header_row1 = [c.value for c in rows[0]]
    header_row2 = [c.value for c in rows[1]]

    # ── detect subject ──
    subject_name: str | None = None
    subject_code: str | None = None

    # Try from row1 merged header
    for val in header_row1[5:]:
        if val and str(val).strip() in SUBJECT_CODE_MAP:
            subject_name = str(val).strip()
            subject_code = SUBJECT_CODE_MAP[subject_name]
            break

    # Fallback: filename
    if not subject_name:
        detected = _detect_subject_from_filename(str(file_path))
        if detected:
            subject_name, subject_code = detected

    if not subject_name or not subject_code:
        return ParsedExamData(warnings=["Cannot detect subject from headers or filename"])

    # ── map row2 columns ──
    # col indices (0-based) for aggregate and question columns
    agg_col_map: dict[str, int] = {}  # e.g. "赋分" -> col_idx
    question_cols: list[tuple[int, str]] = []  # (col_idx, question_name)

    for idx in range(5, len(header_row2)):
        h = header_row2[idx]
        if h is None:
            continue
        h_str = str(h).strip()
        if h_str in _AGGREGATE_HEADERS:
            agg_col_map[h_str] = idx
        elif h_str:
            question_cols.append((idx, h_str))

    # ── build question defs (max_score inferred later) ──
    # Track per-question max observed score for inference.
    q_max: dict[str, float] = {}

    questions_by_name: dict[str, QuestionDef] = {}
    for _, qname in question_cols:
        qtype = _infer_question_type(qname)
        qdef = QuestionDef(
            name=qname,
            question_type=qtype,
            max_score=0,
            max_score_inferred=True,
        )
        questions_by_name[qname] = qdef
        q_max[qname] = 0.0

    # ── parse data rows ──
    students: list[StudentScore] = []
    for row in rows[2:]:
        vals = [c.value for c in row]
        # skip empty rows
        if not vals or all(v is None for v in vals[:5]):
            continue

        school = str(vals[0]).strip() if vals[0] else None
        class_name = str(vals[1]).strip() if vals[1] else None
        student_key = str(vals[2]).strip() if vals[2] else ""
        student_name = str(vals[3]).strip() if vals[3] else ""
        elective = str(vals[4]).strip() if vals[4] else None

        if not student_key and not student_name:
            continue

        converted = _to_float(vals[agg_col_map["赋分"]]) if "赋分" in agg_col_map else None
        raw_total = _to_float(vals[agg_col_map["原始分"]]) if "原始分" in agg_col_map else None
        grade_level = str(vals[agg_col_map["等级"]]).strip() if "等级" in agg_col_map and vals[agg_col_map["等级"]] else None
        class_rank = _to_int(vals[agg_col_map["班名次"]]) if "班名次" in agg_col_map else None
        school_rank = _to_int(vals[agg_col_map["校名次"]]) if "校名次" in agg_col_map else None
        obj_sub = _to_float(vals[agg_col_map["客观题"]]) if "客观题" in agg_col_map else None
        subj_sub = _to_float(vals[agg_col_map["主观题"]]) if "主观题" in agg_col_map else None

        q_scores: dict[str, float] = {}
        for col_idx, qname in question_cols:
            score = _to_float(vals[col_idx]) if col_idx < len(vals) else None
            if score is not None:
                q_scores[qname] = score
                if score > q_max[qname]:
                    q_max[qname] = score

        stu = StudentScore(
            student_key=student_key,
            student_name=student_name,
            class_name=class_name,
            school_name=school,
            elective=elective,
            raw_total=raw_total,
            converted_score=converted,
            grade_level=grade_level,
            class_rank=class_rank,
            school_rank=school_rank,
            objective_subtotal=obj_sub,
            subjective_subtotal=subj_sub,
            question_scores=q_scores,
        )
        students.append(stu)

    # finalize question max_score from observed data
    for qname, qdef in questions_by_name.items():
        qdef.max_score = q_max[qname]

    questions = [questions_by_name[qn] for _, qn in question_cols]

    subject_data = ParsedSubjectData(
        subject_name=subject_name,
        subject_code=subject_code,
        questions=questions,
        students=students,
    )

    wb.close()
    return ParsedExamData(subjects=[subject_data], warnings=warnings)


# ── parse_totals_xlsx ─────────────────────────────────────────────


def parse_totals_xlsx(file_path: str | Path) -> ParsedExamData:
    """Parse a subject-totals xlsx (one column per subject).

    Expected layout:
      Row 1: 考号 | 姓名 | 班级 | 语文 | 语文(赋分) | 数学 | 英语 | ... | 总分 | 班名次 | 校名次
      Row 2+: data
    """
    file_path = Path(file_path)
    wb = openpyxl.load_workbook(str(file_path))
    ws = wb.active
    warnings: list[str] = []

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return ParsedExamData(warnings=["Sheet has fewer than 2 rows"])

    headers = [str(v).strip() if v else "" for v in rows[0]]

    # ── identify column roles ──
    fixed_cols: dict[str, int] = {}  # header -> col_idx
    subject_raw_cols: dict[str, int] = {}  # subject_cn -> col_idx for raw score
    subject_converted_cols: dict[str, int] = {}  # subject_cn -> col_idx for 赋分
    total_col: int | None = None
    class_rank_col: int | None = None
    school_rank_col: int | None = None

    for idx, h in enumerate(headers):
        h_clean = h.strip()

        # Check for 赋分 variant: "语文(赋分)" or "语文（赋分）"
        m_conv = re.match(r"^(.+?)[（(]赋分[）)]$", h_clean)
        if m_conv:
            subj_cn = m_conv.group(1).strip()
            if subj_cn in SUBJECT_CODE_MAP:
                subject_converted_cols[subj_cn] = idx
                continue

        # Skip 等级 columns
        m_grade = re.match(r"^(.+?)[（(]等级[）)]$", h_clean)
        if m_grade:
            continue

        if h_clean in ("总分",):
            total_col = idx
            continue
        if h_clean in ("班名次",):
            class_rank_col = idx
            continue
        if h_clean in ("校名次",):
            school_rank_col = idx
            continue

        lowered = h_clean
        if lowered in _TOTALS_FIXED:
            fixed_cols[lowered] = idx
            continue

        # subject raw score column
        if h_clean in SUBJECT_CODE_MAP:
            subject_raw_cols[h_clean] = idx

    if not subject_raw_cols:
        return ParsedExamData(warnings=["No subject columns detected in headers"])

    # convenience lookups for fixed columns
    def _fc(name: str) -> int | None:
        return fixed_cols.get(name)

    # _fc returns None or int (including 0); cannot use `or` since 0 is falsy.
    key_col = _fc("考号") if _fc("考号") is not None else (_fc("学号") if _fc("学号") is not None else _fc("准考证号"))
    name_col = _fc("姓名")
    class_col = _fc("班级")
    school_col = _fc("学校")
    elective_col = _fc("选科")

    # ── build per-subject structures ──
    subject_map: dict[str, ParsedSubjectData] = {}
    for cn_name, col_idx in subject_raw_cols.items():
        code = SUBJECT_CODE_MAP[cn_name]
        max_score = SUBJECT_MAX_SCORE.get(code, 0)
        qdef = QuestionDef(
            name="__TOTAL__",
            question_type="synthetic",
            max_score=max_score,
        )
        subject_map[cn_name] = ParsedSubjectData(
            subject_name=cn_name,
            subject_code=code,
            questions=[qdef],
            students=[],
        )

    # ── parse data rows ──
    for row_vals in rows[1:]:
        vals = list(row_vals)
        if not vals or all(v is None for v in vals):
            continue

        s_key = str(vals[key_col]).strip() if key_col is not None and vals[key_col] else ""
        s_name = str(vals[name_col]).strip() if name_col is not None and vals[name_col] else ""
        if not s_key and not s_name:
            continue

        s_class = str(vals[class_col]).strip() if class_col is not None and vals[class_col] else None
        s_school = str(vals[school_col]).strip() if school_col is not None and vals[school_col] else None
        s_elective = str(vals[elective_col]).strip() if elective_col is not None and vals[elective_col] else None
        s_class_rank = _to_int(vals[class_rank_col]) if class_rank_col is not None else None
        s_school_rank = _to_int(vals[school_rank_col]) if school_rank_col is not None else None

        for cn_name, col_idx in subject_raw_cols.items():
            raw = _to_float(vals[col_idx]) if col_idx < len(vals) else None
            converted = None
            if cn_name in subject_converted_cols:
                c_idx = subject_converted_cols[cn_name]
                converted = _to_float(vals[c_idx]) if c_idx < len(vals) else None

            stu = StudentScore(
                student_key=s_key,
                student_name=s_name,
                class_name=s_class,
                school_name=s_school,
                elective=s_elective,
                raw_total=raw,
                converted_score=converted,
                class_rank=s_class_rank,
                school_rank=s_school_rank,
                question_scores={"__TOTAL__": raw} if raw is not None else {},
            )
            subject_map[cn_name].students.append(stu)

    subjects = list(subject_map.values())
    wb.close()
    return ParsedExamData(subjects=subjects, warnings=warnings)


# ── parse_zip ─────────────────────────────────────────────────────

MAX_ZIP_FILES = 200
MAX_ZIP_TOTAL_SIZE = 100 * 1024 * 1024  # 100 MB
MAX_ZIP_RATIO = 20


def parse_zip(file_path: str | Path) -> ParsedExamData:
    """Extract xlsx files from a zip and parse them.

    Priority: files matching *小题分* -> parse_question_scores_xlsx
    Fallback: parse_totals_xlsx for remaining xlsx files.
    """
    file_path = Path(file_path)
    warnings: list[str] = []

    with zipfile.ZipFile(str(file_path), "r") as zf:
        # ── safety checks ──
        infos = zf.infolist()

        if len(infos) > MAX_ZIP_FILES:
            return ParsedExamData(warnings=[f"Zip contains {len(infos)} files, exceeds limit {MAX_ZIP_FILES}"])

        total_uncompressed = sum(i.file_size for i in infos)
        if total_uncompressed > MAX_ZIP_TOTAL_SIZE:
            return ParsedExamData(warnings=[f"Zip uncompressed size {total_uncompressed} exceeds limit {MAX_ZIP_TOTAL_SIZE}"])

        total_compressed = sum(i.compress_size for i in infos if i.compress_size > 0)
        if total_compressed > 0 and total_uncompressed / total_compressed > MAX_ZIP_RATIO:
            return ParsedExamData(warnings=[f"Zip compression ratio {total_uncompressed / total_compressed:.1f} exceeds limit {MAX_ZIP_RATIO}"])

        # Filter xlsx entries, reject path traversal and symlinks
        xlsx_entries: list[zipfile.ZipInfo] = []
        for info in infos:
            if info.is_dir():
                continue
            # path traversal check
            if ".." in info.filename or info.filename.startswith("/"):
                warnings.append(f"Skipping suspicious path: {info.filename}")
                continue
            # symlink check (external_attr bit)
            if (info.external_attr >> 16) & 0o170000 == 0o120000:
                warnings.append(f"Skipping symlink: {info.filename}")
                continue
            if info.filename.lower().endswith(".xlsx") and not info.filename.startswith("__MACOSX"):
                xlsx_entries.append(info)

        if not xlsx_entries:
            return ParsedExamData(warnings=["No xlsx files found in zip"])

        # ── extract and parse ──
        all_subjects: list[ParsedSubjectData] = []

        with tempfile.TemporaryDirectory() as tmpdir:
            for entry in xlsx_entries:
                safe_name = os.path.basename(entry.filename)
                target = os.path.join(tmpdir, safe_name)
                with open(target, "wb") as f:
                    f.write(zf.read(entry.filename))

                # Decide parser
                if "小题分" in safe_name:
                    result = parse_question_scores_xlsx(target)
                else:
                    result = parse_totals_xlsx(target)

                all_subjects.extend(result.subjects)
                warnings.extend(result.warnings)

    return ParsedExamData(subjects=all_subjects, warnings=warnings)
