"""教师管理路由 — 列表/创建/导入/导出/编辑/删除。"""
import io
import logging

from fastapi import APIRouter, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from edu_cloud.database import get_db
from edu_cloud.api.deps import get_current_user
from edu_cloud.models.user import User
from edu_cloud.models.user_role import UserRole
from edu_cloud.modules.student.models import Class
from edu_cloud.services.exceptions import NotFoundError, ValidationError

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1", tags=["teachers"])

TEACHER_ROLES = {
    "subject_teacher", "homeroom_teacher", "teaching_research_leader",
    "grade_leader", "lesson_prep_leader",
}
ALL_SCHOOL_ROLES = TEACHER_ROLES | {
    "principal", "academic_director", "district_admin",
}

_ROLE_LABELS = {
    "platform_admin": "平台管理员", "district_admin": "区管理员",
    "principal": "校长", "academic_director": "教务主任",
    "teaching_research_leader": "教研组长", "grade_leader": "年级组长",
    "lesson_prep_leader": "备课组长", "homeroom_teacher": "班主任",
    "subject_teacher": "科任教师",
}


class TeacherCreate(BaseModel):
    username: str
    display_name: str
    password: str = "123456"
    roles: list[str] = ["subject_teacher"]
    phone: str | None = None
    email: str | None = None
    employee_id: str | None = None
    gender: str | None = None
    id_card: str | None = None
    title: str | None = None
    hire_date: str | None = None
    education: str | None = None
    university: str | None = None
    office_phone: str | None = None
    notes: str | None = None
    subject_codes: list[str] | None = None
    class_ids: list[str] | None = None


class TeacherUpdate(BaseModel):
    display_name: str | None = None
    phone: str | None = None
    email: str | None = None
    employee_id: str | None = None
    gender: str | None = None
    id_card: str | None = None
    title: str | None = None
    hire_date: str | None = None
    education: str | None = None
    university: str | None = None
    office_phone: str | None = None
    notes: str | None = None
    is_active: bool | None = None


def _teacher_response(user: User, roles: list[UserRole]) -> dict:
    return {
        "id": user.id,
        "username": user.username,
        "display_name": user.display_name,
        "phone": user.phone,
        "email": user.email,
        "employee_id": user.employee_id,
        "gender": user.gender,
        "id_card": user.id_card,
        "title": user.title,
        "hire_date": str(user.hire_date) if user.hire_date else None,
        "education": user.education,
        "university": user.university,
        "office_phone": user.office_phone,
        "notes": user.notes,
        "is_active": user.is_active,
        "roles": [
            {
                "id": r.id, "role": r.role,
                "role_label": _ROLE_LABELS.get(r.role, r.role),
                "subject_codes": r.subject_codes,
                "class_ids": r.class_ids,
            }
            for r in roles
        ],
    }


@router.get("/teachers")
async def list_teachers(
    q: str | None = None,
    school_id: str | None = None,
    db: AsyncSession = Depends(get_db),
    current: dict = Depends(get_current_user),
):
    role = current["current_role"]
    target_school = school_id or role.school_id
    if not target_school:
        # platform_admin 无 school_id 时，取第一个学校
        from edu_cloud.modules.school.models import School
        first = await db.execute(select(School).limit(1))
        s = first.scalar_one_or_none()
        target_school = s.id if s else None
    if not target_school:
        return []

    stmt = (
        select(User)
        .join(UserRole, UserRole.user_id == User.id)
        .where(UserRole.school_id == target_school, User.is_active.is_(True))
        .distinct()
    )
    if q:
        stmt = stmt.where(User.display_name.contains(q) | User.username.contains(q))
    result = await db.execute(stmt)
    users = list(result.scalars().all())

    roles_result = await db.execute(
        select(UserRole).where(
            UserRole.school_id == target_school,
            UserRole.user_id.in_([u.id for u in users]),
        )
    )
    all_roles = list(roles_result.scalars().all())
    roles_by_user = {}
    for r in all_roles:
        roles_by_user.setdefault(r.user_id, []).append(r)

    return [_teacher_response(u, roles_by_user.get(u.id, [])) for u in users]


@router.post("/teachers", status_code=201)
async def create_teacher(
    req: TeacherCreate,
    db: AsyncSession = Depends(get_db),
    current: dict = Depends(get_current_user),
):
    school_id = current["current_role"].school_id or (
        req.school_id if hasattr(req, 'school_id') else None
    )
    existing = await db.execute(select(User).where(User.username == req.username))
    if existing.scalar_one_or_none():
        raise ValidationError(f"用户名 {req.username} 已存在")
    for r in req.roles:
        if r not in ALL_SCHOOL_ROLES:
            raise ValidationError(f"角色 {r} 不合法")

    import datetime as _dt
    user = User(
        username=req.username, display_name=req.display_name,
        phone=req.phone, email=req.email,
        employee_id=req.employee_id, gender=req.gender, id_card=req.id_card,
        title=req.title,
        hire_date=_dt.date.fromisoformat(req.hire_date) if req.hire_date else None,
        education=req.education, university=req.university,
        office_phone=req.office_phone, notes=req.notes,
    )
    user.set_password(req.password)
    db.add(user)
    await db.flush()

    created_roles = []
    for i, role_name in enumerate(req.roles):
        ur = UserRole(
            user_id=user.id, role=role_name, school_id=school_id,
            subject_codes=req.subject_codes, class_ids=req.class_ids,
            is_primary=(i == 0),
        )
        db.add(ur)
        created_roles.append(ur)
    await db.commit()
    await db.refresh(user)
    return _teacher_response(user, created_roles)


@router.patch("/teachers/{user_id}")
async def update_teacher(
    user_id: str,
    req: TeacherUpdate,
    db: AsyncSession = Depends(get_db),
    current: dict = Depends(get_current_user),
):
    school_id = current["current_role"].school_id
    role_check = await db.execute(
        select(UserRole).where(UserRole.user_id == user_id, UserRole.school_id == school_id)
    )
    if not role_check.scalars().first():
        raise NotFoundError("教师不存在")
    user_result = await db.execute(select(User).where(User.id == user_id))
    user = user_result.scalar_one_or_none()
    if not user:
        raise NotFoundError("用户不存在")
    import datetime as _dt
    for f in ["display_name", "phone", "email", "employee_id", "gender",
              "id_card", "title", "education", "university", "office_phone",
              "notes", "is_active"]:
        v = getattr(req, f, None)
        if v is not None:
            setattr(user, f, v)
    if req.hire_date is not None:
        user.hire_date = _dt.date.fromisoformat(req.hire_date) if req.hire_date else None
    await db.commit()
    await db.refresh(user)
    roles_result = await db.execute(
        select(UserRole).where(UserRole.user_id == user_id, UserRole.school_id == school_id)
    )
    return _teacher_response(user, list(roles_result.scalars().all()))


@router.delete("/teachers/{user_id}", status_code=204)
async def delete_teacher(
    user_id: str,
    db: AsyncSession = Depends(get_db),
    current: dict = Depends(get_current_user),
):
    school_id = current["current_role"].school_id
    roles_result = await db.execute(
        select(UserRole).where(UserRole.user_id == user_id, UserRole.school_id == school_id)
    )
    roles = list(roles_result.scalars().all())
    if not roles:
        raise NotFoundError("教师不存在")
    for r in roles:
        await db.delete(r)
    remaining = await db.execute(
        select(func.count()).select_from(UserRole).where(UserRole.user_id == user_id)
    )
    if remaining.scalar() == 0:
        user_result = await db.execute(select(User).where(User.id == user_id))
        user = user_result.scalar_one_or_none()
        if user:
            await db.delete(user)
    await db.commit()


@router.get("/teachers/export")
async def export_teachers(
    template: str | None = None,
    school_id: str | None = None,
    db: AsyncSession = Depends(get_db),
    current: dict = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.comments import Comment

    target_school = school_id or current["current_role"].school_id
    if not target_school:
        from edu_cloud.modules.school.models import School
        first = await db.execute(select(School).limit(1))
        s = first.scalar_one_or_none()
        target_school = s.id if s else None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "教师名单"
    headers = [
        "姓名", "工号", "任教学科", "年级", "任教班级",
        "角色", "手机", "性别", "职称", "入职日期",
        "学历", "毕业院校", "邮箱", "办公电话", "身份证号", "备注",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    widths = [12, 18, 14, 10, 28, 18, 15, 8, 12, 14, 10, 20, 22, 15, 22, 20]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    classes_result = await db.execute(select(Class).where(Class.school_id == target_school))
    all_classes = list(classes_result.scalars().all())
    class_id_to_name = {c.id: c.name for c in all_classes}
    class_name_to_id = {c.name: c.id for c in all_classes}

    _SUBJECT_LABELS = {
        "YW": "语文", "SX": "数学", "YY": "英语", "WL": "物理", "HX": "化学",
        "SW": "生物", "ZZ": "政治", "LS": "历史", "DL": "地理", "TY": "体育",
        "YS": "音乐", "MS": "美术", "XX": "信息技术",
    }

    if template == "1":
        # 列序: A姓名 B工号 C学科 D年级 E班级 F角色 G手机 H性别 I职称 J入职 K学历 L院校 M邮箱 N办公电话 O身份证 P备注
        ws["H1"].comment = Comment("填写：男 或 女", "系统")
        ws["I1"].comment = Comment("如：一级教师、高级教师、特级教师", "系统")
        ws["J1"].comment = Comment("格式：2020-09-01", "系统")
        ws["K1"].comment = Comment("如：本科、硕士、博士", "系统")
        role_labels = "\n".join(f"{v}（{k}）" for k, v in _ROLE_LABELS.items() if k in ALL_SCHOOL_ROLES)
        ws["F1"].comment = Comment(f"多角色用逗号分隔，填中文或代码均可：\n{role_labels}", "系统")
        subject_hint = "\n".join(f"{v}（{k}）" for k, v in _SUBJECT_LABELS.items())
        ws["C1"].comment = Comment(f"多学科用逗号分隔，填中文或代码均可：\n{subject_hint}", "系统")
        class_names = [c.name for c in all_classes[:30]]
        ws["E1"].comment = Comment(f"多班级用逗号分隔。当前班级：\n" + "\n".join(class_names), "系统")
        ws["D1"].comment = Comment("从班级名自动推导，可留空", "系统")

        example_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        example_class = all_classes[0].name if all_classes else "2301"
        example_grade = all_classes[0].grade if all_classes else "高三"
        ws.append(["张三（示例）", "T2026001", "语文", example_grade, example_class,
                   "班主任,科任教师", "13800138000", "男",
                   "一级教师", "2020-09-01", "硕士", "北京师范大学", "", "", "", ""])
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=2, column=col_idx).fill = example_fill

        ws2 = wb.create_sheet("班级参考")
        ws2.append(["班级名称", "年级"])
        for c in all_classes:
            ws2.append([c.name, c.grade or ""])

        ws3 = wb.create_sheet("角色与学科参考")
        ws3.append(["角色代码", "角色名称"])
        for k, v in _ROLE_LABELS.items():
            if k in ALL_SCHOOL_ROLES:
                ws3.append([k, v])
        ws3.append([])
        ws3.append(["学科代码", "学科名称"])
        for k, v in _SUBJECT_LABELS.items():
            ws3.append([k, v])
        filename = "teachers_template.xlsx"
    else:
        stmt = (
            select(User).join(UserRole, UserRole.user_id == User.id)
            .where(UserRole.school_id == target_school).distinct()
        )
        result = await db.execute(stmt)
        users = list(result.scalars().all())
        roles_result = await db.execute(
            select(UserRole).where(UserRole.school_id == target_school, UserRole.user_id.in_([u.id for u in users]))
        )
        class_id_to_grade = {c.id: (c.grade or "") for c in all_classes}
        roles_by_user = {}
        for r in roles_result.scalars().all():
            roles_by_user.setdefault(r.user_id, []).append(r)
        for u in users:
            user_roles = roles_by_user.get(u.id, [])
            role_parts = []
            subjects = set()
            teach_classes = set()
            grades = set()
            for r in user_roles:
                for sc in (r.subject_codes or []):
                    subjects.add(_SUBJECT_LABELS.get(sc, sc))
                if r.role == 'subject_teacher':
                    for cid in (r.class_ids or []):
                        cname = class_id_to_name.get(cid, cid)
                        teach_classes.add(cname)
                        g = class_id_to_grade.get(cid)
                        if g:
                            grades.add(g)
                if r.role == 'homeroom_teacher':
                    hr_cid = (r.class_ids or [None])[0]
                    hr_name = class_id_to_name.get(hr_cid, '') if hr_cid else ''
                    label = _ROLE_LABELS.get(r.role, r.role)
                    role_parts.append(f"{label}({hr_name})" if hr_name else label)
                elif r.role != 'subject_teacher':
                    role_parts.append(_ROLE_LABELS.get(r.role, r.role))
            if any(r.role == 'subject_teacher' for r in user_roles):
                role_parts.append('科任教师')
            role_str = ",".join(role_parts)
            ws.append([
                u.display_name, u.employee_id or u.username,
                ",".join(sorted(subjects)), ",".join(sorted(grades)),
                ",".join(sorted(teach_classes)), role_str,
                u.phone or "", u.gender or "", u.title or "",
                str(u.hire_date) if u.hire_date else "",
                u.education or "", u.university or "",
                u.email or "", u.office_phone or "", u.id_card or "",
                u.notes or "",
            ])
        filename = "teachers.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/teachers/import", status_code=201)
async def import_teachers(
    file: UploadFile = File(...),
    role: str = Form("subject_teacher"),
    db: AsyncSession = Depends(get_db),
    current: dict = Depends(get_current_user),
):
    school_id = current["current_role"].school_id
    if role not in ALL_SCHOOL_ROLES:
        raise ValidationError(f"角色 {role} 不合法")

    import openpyxl
    import datetime as _dt

    _SUBJECT_LABELS = {
        "YW": "语文", "SX": "数学", "YY": "英语", "WL": "物理", "HX": "化学",
        "SW": "生物", "ZZ": "政治", "LS": "历史", "DL": "地理", "TY": "体育",
        "YS": "音乐", "MS": "美术", "XX": "信息技术",
    }
    _LABEL_TO_SUBJECT = {v: k for k, v in _SUBJECT_LABELS.items()}
    _LABEL_TO_ROLE = {v: k for k, v in _ROLE_LABELS.items()}

    def _resolve_subject(s):
        s = s.strip()
        return _LABEL_TO_SUBJECT.get(s, s) if s not in _SUBJECT_LABELS else s

    def _resolve_role(r):
        r = r.strip()
        return _LABEL_TO_ROLE.get(r, r) if r not in ALL_SCHOOL_ROLES else r

    content = await file.read()
    filename = file.filename or ""
    if not filename.endswith((".xlsx", ".xls")):
        raise ValidationError("仅支持 .xlsx/.xls 文件")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValidationError("Excel 为空或仅含表头")

    header = [str(h).strip() if h else "" for h in rows[0]]

    def _find(keywords):
        return next((i for i, h in enumerate(header) if any(k in h for k in keywords)), None)

    name_col = _find(["姓名"])
    username_col = _find(["工号", "账号", "用户名"])
    phone_col = _find(["手机", "电话"])
    email_col = _find(["邮箱"])
    gender_col = _find(["性别"])
    id_card_col = _find(["身份证", "证件"])
    title_col = _find(["职称"])
    hire_col = _find(["入职"])
    edu_col = _find(["学历"])
    uni_col = _find(["毕业", "院校"])
    office_col = _find(["办公电话"])
    role_col = _find(["角色"])
    subject_col = _find(["学科", "任教学科"])
    class_col = _find(["班级", "任教班级"])
    notes_col = _find(["备注"])

    if name_col is None:
        raise ValidationError("Excel 需包含「姓名」列")

    classes_result = await db.execute(select(Class).where(Class.school_id == school_id))
    class_name_to_id = {c.name: c.id for c in classes_result.scalars().all()}

    def _cell(row, col):
        if col is None or col >= len(row) or row[col] is None:
            return None
        return str(row[col]).strip() or None

    created = 0
    updated = 0
    skipped = 0
    for row in rows[1:]:
        display_name = _cell(row, name_col)
        if not display_name or "示例" in display_name:
            continue

        username = _cell(row, username_col) or f"t_{display_name}"
        raw_role = _cell(row, role_col) or role
        row_roles = [_resolve_role(r) for r in raw_role.replace("，", ",").split(",") if r.strip()]
        row_roles = [r for r in row_roles if r in ALL_SCHOOL_ROLES] or [role]

        existing = await db.execute(select(User).where(User.username == username))
        user = existing.scalar_one_or_none()
        if user:
            changed = False
            for attr, col in [
                ("phone", phone_col), ("email", email_col), ("gender", gender_col),
                ("id_card", id_card_col), ("title", title_col), ("education", edu_col),
                ("university", uni_col), ("office_phone", office_col), ("notes", notes_col),
                ("employee_id", username_col),
            ]:
                v = _cell(row, col)
                if v and not getattr(user, attr, None):
                    setattr(user, attr, v)
                    changed = True
            hd = _cell(row, hire_col)
            if hd and not user.hire_date:
                try:
                    user.hire_date = _dt.date.fromisoformat(hd)
                    changed = True
                except ValueError:
                    pass
            # 更新 UserRole 的 subject_codes / class_ids
            role_result = await db.execute(
                select(UserRole).where(UserRole.user_id == user.id, UserRole.school_id == school_id)
            )
            user_role = role_result.scalars().first()
            if user_role:
                sc = _cell(row, subject_col) if subject_col is not None else None
                if sc and not user_role.subject_codes:
                    user_role.subject_codes = [_resolve_subject(s) for s in sc.replace("，", ",").split(",") if s.strip()]
                    changed = True
                cc = _cell(row, class_col) if class_col is not None else None
                if cc and not user_role.class_ids:
                    cids = [class_name_to_id[cn.strip()] for cn in cc.replace("，", ",").split(",")
                            if cn.strip() in class_name_to_id]
                    if cids:
                        user_role.class_ids = cids
                        changed = True
            if changed:
                updated += 1
            else:
                skipped += 1
            continue

        hire_date = None
        hd = _cell(row, hire_col)
        if hd:
            try:
                hire_date = _dt.date.fromisoformat(hd)
            except ValueError:
                pass

        user = User(
            username=username, display_name=display_name,
            phone=_cell(row, phone_col), email=_cell(row, email_col),
            employee_id=_cell(row, username_col), gender=_cell(row, gender_col),
            id_card=_cell(row, id_card_col), title=_cell(row, title_col),
            hire_date=hire_date, education=_cell(row, edu_col),
            university=_cell(row, uni_col), office_phone=_cell(row, office_col),
            notes=_cell(row, notes_col),
        )
        user.set_password("123456")
        db.add(user)
        await db.flush()

        subject_codes = None
        if subject_col is not None:
            sc = _cell(row, subject_col)
            if sc:
                subject_codes = [_resolve_subject(s) for s in sc.replace("，", ",").split(",") if s.strip()]

        class_ids = None
        if class_col is not None:
            cc = _cell(row, class_col)
            if cc:
                class_ids = []
                for cn in cc.replace("，", ",").split(","):
                    cn = cn.strip()
                    cid = class_name_to_id.get(cn)
                    if cid:
                        class_ids.append(cid)
                class_ids = class_ids or None

        for i, rn in enumerate(row_roles):
            ur = UserRole(
                user_id=user.id, role=rn, school_id=school_id,
                subject_codes=subject_codes, class_ids=class_ids,
                is_primary=(i == 0),
            )
            db.add(ur)
        created += 1

    await db.commit()
    logger.info("import_teachers: school=%s, created=%d, updated=%d, skipped=%d",
                school_id, created, updated, skipped)
    return {"created": created, "updated": updated, "skipped": skipped}
