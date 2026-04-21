"""Class / Student 路由 — 从 exam-ai 迁入。"""
import io
import logging

from fastapi import APIRouter, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from edu_cloud.database import get_db
from edu_cloud.api.deps import get_current_user
from edu_cloud.api.permissions import get_visible_class_ids
from edu_cloud.modules.student import service as student_service
from edu_cloud.modules.student.models import Class
from edu_cloud.models.subject_selection import SubjectSelection

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1", tags=["students"])


class StudentCreate(BaseModel):
    name: str
    student_number: str
    class_id: str | None = None
    grade: str | None = None
    gender: str | None = None
    id_card: str | None = None
    selection_id: str | None = None


class StudentUpdate(BaseModel):
    name: str | None = None
    student_number: str | None = None
    class_id: str | None = None
    grade: str | None = None
    gender: str | None = None
    id_card: str | None = None
    selection_id: str | None = None


def _class_response(c) -> dict:
    return {"id": c.id, "name": c.name, "grade": c.grade,
            "head_teacher_id": c.head_teacher_id, "school_id": c.school_id}


def _student_response(s) -> dict:
    return {"id": s.id, "name": s.name, "student_number": s.student_number,
            "class_id": s.class_id, "school_id": s.school_id,
            "gender": s.gender, "grade": s.grade,
            "id_card": s.id_card, "selection_id": s.selection_id}


@router.get("/classes")
async def list_classes(
    grade: str | None = None,
    db: AsyncSession = Depends(get_db),
    current: dict = Depends(get_current_user),
):
    role = current["current_role"]
    classes = await student_service.list_classes(
        db, school_id=role.school_id, grade=grade,
        visible_class_ids=get_visible_class_ids(role),
    )
    return [_class_response(c) for c in classes]


@router.get("/grades")
async def list_grades(
    db: AsyncSession = Depends(get_db),
    current: dict = Depends(get_current_user),
):
    role = current["current_role"]
    classes = await student_service.list_classes(
        db, school_id=role.school_id,
        visible_class_ids=get_visible_class_ids(role),
    )
    grades = sorted(set(c.grade for c in classes if c.grade))
    return [{"name": g} for g in grades]


@router.get("/students")
async def list_students(
    class_id: str | None = None,
    grade: str | None = None,
    selection_id: str | None = None,
    subject_code: str | None = None,
    q: str | None = None,
    db: AsyncSession = Depends(get_db),
    current: dict = Depends(get_current_user),
):
    role = current["current_role"]
    if grade and not class_id:
        grade_classes = await student_service.list_classes(
            db, school_id=role.school_id, grade=grade,
            visible_class_ids=get_visible_class_ids(role),
        )
        grade_class_ids = [c.id for c in grade_classes]
        all_students = []
        for cid in grade_class_ids:
            batch = await student_service.list_students(
                db, school_id=role.school_id, class_id=cid,
                selection_id=selection_id, subject_code=subject_code,
                visible_class_ids=get_visible_class_ids(role),
            )
            all_students.extend(batch)
        return [_student_response(s) for s in all_students]
    if q:
        students = await student_service.search_students(
            db, school_id=role.school_id, query=q,
            visible_class_ids=get_visible_class_ids(role),
        )
    else:
        students = await student_service.list_students(
            db, school_id=role.school_id, class_id=class_id,
            selection_id=selection_id, subject_code=subject_code,
            visible_class_ids=get_visible_class_ids(role),
        )
    return [_student_response(s) for s in students]


@router.post("/students", status_code=201)
async def create_student(
    req: StudentCreate,
    db: AsyncSession = Depends(get_db),
    current: dict = Depends(get_current_user),
):
    role = current["current_role"]
    student = await student_service.create_student(
        db, school_id=role.school_id,
        name=req.name, student_number=req.student_number,
        class_id=req.class_id, grade=req.grade, gender=req.gender,
        id_card=req.id_card, selection_id=req.selection_id,
    )
    return _student_response(student)


@router.patch("/students/{student_id}")
async def update_student(
    student_id: str,
    req: StudentUpdate,
    db: AsyncSession = Depends(get_db),
    current: dict = Depends(get_current_user),
):
    role = current["current_role"]
    student = await student_service.update_student(
        db, student_id=student_id, school_id=role.school_id,
        name=req.name, student_number=req.student_number,
        class_id=req.class_id, grade=req.grade, gender=req.gender,
        id_card=req.id_card, selection_id=req.selection_id,
    )
    return _student_response(student)


@router.delete("/students/{student_id}", status_code=204)
async def delete_student(
    student_id: str,
    db: AsyncSession = Depends(get_db),
    current: dict = Depends(get_current_user),
):
    role = current["current_role"]
    await student_service.delete_student(
        db, student_id=student_id, school_id=role.school_id,
    )


@router.post("/students/import", status_code=201)
async def import_students(
    file: UploadFile = File(...),
    class_id: str = Form(""),
    grade: str = Form(""),
    db: AsyncSession = Depends(get_db),
    current: dict = Depends(get_current_user),
):
    role = current["current_role"]
    content = await file.read()
    return await student_service.import_students(
        db, school_id=role.school_id, class_id=class_id, grade=grade,
        file_content=content, filename=file.filename or "",
        visible_class_ids=get_visible_class_ids(role),
    )


@router.get("/students/export")
async def export_students(
    class_id: str | None = None,
    template: str | None = None,
    db: AsyncSession = Depends(get_db),
    current: dict = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    role = current["current_role"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生名单"
    headers = ["姓名", "学号/准考证号", "班级", "年级", "性别", "身份证号", "选课组合"]
    ws.append(headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 20

    classes_result = await db.execute(
        select(Class).where(Class.school_id == role.school_id)
    )
    all_classes = list(classes_result.scalars().all())
    class_map = {c.id: c.name for c in all_classes}

    if template == "1":
        from openpyxl.comments import Comment

        sel_result = await db.execute(
            select(SubjectSelection).where(
                SubjectSelection.school_id == role.school_id,
                SubjectSelection.is_active.is_(True),
            )
        )
        all_sels = list(sel_result.scalars().all())
        sel_names = [s.name for s in all_sels]

        class_names = [c.name for c in all_classes]
        if class_names:
            ws["C1"].comment = Comment(
                f"请填写以下班级名称之一：\n" + "\n".join(class_names),
                "系统",
            )
        if sel_names:
            ws["G1"].comment = Comment(
                f"请填写以下选课组合名称之一：\n" + "\n".join(sel_names),
                "系统",
            )

        example_class = class_names[0] if class_names else "高一(1)班"
        example_sel = sel_names[0] if sel_names else ""
        ws.append(["张三（示例）", "2026001", example_class, "高一", "男", "", example_sel])
        example_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=2, column=col_idx).fill = example_fill

        ws2 = wb.create_sheet("班级参考")
        ws2.append(["班级名称", "年级"])
        for c in all_classes:
            ws2.append([c.name, c.grade or ""])

        ws3 = wb.create_sheet("选课组合参考")
        ws3.append(["组合名称", "科目代码"])
        for s in all_sels:
            ws3.append([s.name, "/".join(s.subject_codes or [])])
        filename = "students_template.xlsx"
    else:
        students = await student_service.list_students(
            db, school_id=role.school_id, class_id=class_id,
            visible_class_ids=get_visible_class_ids(role),
        )
        sel_result = await db.execute(
            select(SubjectSelection).where(SubjectSelection.school_id == role.school_id)
        )
        sel_map = {s.id: s.name for s in sel_result.scalars().all()}
        for s in students:
            ws.append([
                s.name,
                s.student_number,
                class_map.get(s.class_id, ""),
                s.grade or "",
                s.gender or "",
                s.id_card or "",
                sel_map.get(s.selection_id, ""),
            ])
        filename = "students.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
