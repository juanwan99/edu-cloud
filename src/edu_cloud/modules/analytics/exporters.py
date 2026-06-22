"""Phase 2-A 报告导出：年级学科报告 PDF / XLSX。

PDF：reportlab Platypus（避开 playwright 依赖，生产环境 CJK 字体走 Noto）
XLSX：openpyxl 多 sheet（总览 / 班级对比 / 题目分析）
"""
from __future__ import annotations

import io
import logging
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from edu_cloud.modules.analytics import service as analytics_service
from edu_cloud.services.analytics_workflow import Exam, Subject
from edu_cloud.services.exceptions import NotFoundError

logger = logging.getLogger(__name__)


# 优先级：项目静态字体 → 系统 Noto → 默认（无 CJK）
_FONT_CANDIDATES = (
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/opentype/noto/NotoSerifCJK-Regular.ttc",
    "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
    "/Library/Fonts/PingFang.ttc",
    "C:/Windows/Fonts/msyh.ttc",
)

_REGISTERED_FONT: str | None = None


def _ensure_cjk_font() -> str:
    """注册 CJK 字体，返回字体名（首次调用时检测）。失败回退到 Helvetica。"""
    global _REGISTERED_FONT
    if _REGISTERED_FONT is not None:
        return _REGISTERED_FONT

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    for path in _FONT_CANDIDATES:
        if Path(path).exists():
            try:
                pdfmetrics.registerFont(TTFont("CJK", path, subfontIndex=0))
                _REGISTERED_FONT = "CJK"
                logger.info("registered CJK font: %s", path)
                return _REGISTERED_FONT
            except Exception as e:
                logger.warning("failed to register %s: %s", path, e)
                continue

    _REGISTERED_FONT = "Helvetica"
    logger.warning("no CJK font found; PDF will not render Chinese characters correctly")
    return _REGISTERED_FONT


async def build_student_subject_report(
    db: AsyncSession,
    *,
    student_id: str,
    exam_id: str,
    subject_id: str,
    school_id: str,
    visible_subject_codes: list[str] | None,
    visible_class_ids: list[str] | None,
) -> dict:
    """聚合个人学科报告所需数据。

    数据：学生信息 / 总分对比 / 各题得失分 / 薄弱题 top3 (<60%)
    数据源：StudentAnswer + GradingResult.final_score（grading_results 优先，
    缺失时回落到 student_answers.score）+ Class.name + Question.
    """
    from edu_cloud.services.analytics_workflow import Question
    from edu_cloud.services.analytics_workflow import StudentAnswer
    from edu_cloud.services.analytics_workflow import GradingResult
    from edu_cloud.services.analytics_workflow import Class, Student

    exam = (await db.execute(
        select(Exam).where(Exam.id == exam_id, Exam.school_id == school_id)
    )).scalar_one_or_none()
    if not exam:
        raise NotFoundError("Exam not found")

    subject = (await db.execute(
        select(Subject).where(
            Subject.id == subject_id,
            Subject.exam_id == exam_id,
            Subject.school_id == school_id,
        )
    )).scalar_one_or_none()
    if not subject:
        raise NotFoundError("Subject not found")

    student = (await db.execute(
        select(Student).where(
            Student.id == student_id, Student.school_id == school_id,
        )
    )).scalar_one_or_none()
    if not student:
        raise NotFoundError("Student not found")

    if visible_subject_codes is not None and subject.code not in visible_subject_codes:
        from edu_cloud.services.exceptions import PermissionDeniedError
        raise PermissionDeniedError("无权访问该科目")
    if visible_class_ids is not None and student.class_id not in visible_class_ids:
        from edu_cloud.services.exceptions import PermissionDeniedError
        raise PermissionDeniedError("无权访问该学生")

    class_obj = None
    if student.class_id:
        class_obj = (await db.execute(
            select(Class).where(Class.id == student.class_id)
        )).scalar_one_or_none()

    # 题目总览
    questions = (await db.execute(
        select(Question).where(
            Question.subject_id == subject_id, Question.school_id == school_id,
        )
    )).scalars().all()
    q_by_id = {q.id: q for q in questions}

    # 该学生的所有 StudentAnswer + GradingResult.final_score 关联
    answers = (await db.execute(
        select(StudentAnswer).where(
            StudentAnswer.subject_id == subject_id,
            StudentAnswer.student_id == student_id,
            StudentAnswer.school_id == school_id,
        )
    )).scalars().all()
    answer_ids = [a.id for a in answers]

    final_by_answer: dict[str, float] = {}
    if answer_ids:
        grs = (await db.execute(
            select(GradingResult).where(
                GradingResult.answer_id.in_(answer_ids),
                GradingResult.school_id == school_id,
            )
        )).scalars().all()
        final_by_answer = {gr.answer_id: gr.final_score for gr in grs if gr.final_score is not None}

    question_lines = []
    total_score = 0.0
    total_max = 0.0
    for a in answers:
        q = q_by_id.get(a.question_id)
        if not q:
            continue
        # 优先用 GradingResult.final_score；否则回落到 StudentAnswer.score
        score = final_by_answer.get(a.id)
        if score is None:
            score = a.score
        score = float(score) if score is not None else 0.0
        max_score = float(q.max_score or 0.0)
        rate = (score / max_score) if max_score > 0 else 0.0
        question_lines.append({
            "question_id": q.id,
            "question_name": q.name,
            "question_type": q.question_type,
            "score": round(score, 2),
            "max_score": max_score,
            "score_rate": round(rate, 4),
        })
        total_score += score
        total_max += max_score
    question_lines.sort(key=lambda x: x["question_name"])

    # 班级均分对比（只看该 subject 的有效分）
    from edu_cloud.services.effective_scores import get_effective_scores
    all_scores = await get_effective_scores(db, subject_id, school_id)
    student_totals: dict[str, float] = {}
    for s in all_scores:
        student_totals[s["student_id"]] = student_totals.get(s["student_id"], 0.0) + s["effective_score"]

    # 班级均分（仅当前学生班级）
    class_avg = None
    class_size = 0
    if class_obj:
        class_member_ids = (await db.execute(
            select(Student.id).where(Student.class_id == class_obj.id)
        )).scalars().all()
        class_member_set = set(class_member_ids)
        class_totals = [t for sid, t in student_totals.items() if sid in class_member_set]
        if class_totals:
            class_avg = round(sum(class_totals) / len(class_totals), 2)
            class_size = len(class_totals)

    # 薄弱题 top3 (<60%)
    weakness = sorted(
        [q for q in question_lines if q["score_rate"] < 0.6],
        key=lambda q: q["score_rate"],
    )[:3]

    return {
        "student": {
            "id": student.id,
            "name": student.name,
            "student_number": student.student_number,
            "class_name": class_obj.name if class_obj else None,
        },
        "exam": {"id": exam.id, "name": exam.name},
        "subject": {"id": subject.id, "name": subject.name, "code": subject.code},
        "totals": {
            "score": round(total_score, 2),
            "max_score": round(total_max, 2),
            "score_rate": round((total_score / total_max), 4) if total_max > 0 else 0.0,
            "class_avg": class_avg,
            "class_size": class_size,
        },
        "questions": question_lines,
        "weakness": weakness,
    }


def render_student_subject_report_pdf(report: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph

    font = _ensure_cjk_font()
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="CN", parent=styles["Normal"], fontName=font, fontSize=10, leading=14))
    styles.add(ParagraphStyle(
        name="CN-H1", parent=styles["Heading1"], fontName=font, fontSize=18, leading=22, spaceAfter=8,
    ))
    styles.add(ParagraphStyle(
        name="CN-H2", parent=styles["Heading2"], fontName=font, fontSize=13, leading=18,
        spaceBefore=10, spaceAfter=6,
    ))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=20 * mm, rightMargin=20 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
    )
    flow = []
    title = (
        f"{report['student']['name']} · {report['exam']['name']} · "
        f"{report['subject']['name']} 个人报告"
    )
    flow.append(Paragraph(title, styles["CN-H1"]))

    # 1. 学生信息 + 总分
    flow.append(Paragraph("一、个人成绩", styles["CN-H2"]))
    s = report["student"]
    t = report["totals"]
    info_rows = [
        ["姓名", s.get("name", "")],
        ["学号", s.get("student_number", "")],
        ["班级", s.get("class_name") or "—"],
        ["总分 / 满分", f"{_fmt(t['score'])} / {_fmt(t['max_score'])}"],
        ["得分率", _pct(t.get("score_rate"))],
        ["班级均分", _fmt(t.get("class_avg")) + (f"（{t.get('class_size')}人）" if t.get("class_size") else "")],
    ]
    flow.append(_table(info_rows, font, col_widths=[40 * mm, 80 * mm]))

    # 2. 各题得失分
    flow.append(Paragraph("二、各题得失分", styles["CN-H2"]))
    q_rows = [["题号", "题型", "得分 / 满分", "得分率"]]
    for q in report["questions"]:
        q_rows.append([
            q["question_name"], q["question_type"],
            f"{_fmt(q['score'])} / {_fmt(q['max_score'])}",
            _pct(q["score_rate"]),
        ])
    flow.append(_table(q_rows, font, col_widths=[35 * mm, 30 * mm, 50 * mm, 25 * mm], header=True))

    # 3. 薄弱题 top3
    flow.append(Paragraph("三、薄弱题（得分率 <60%，top 3）", styles["CN-H2"]))
    if report["weakness"]:
        w_rows = [["题号", "题型", "得分 / 满分", "得分率"]]
        for q in report["weakness"]:
            w_rows.append([
                q["question_name"], q["question_type"],
                f"{_fmt(q['score'])} / {_fmt(q['max_score'])}",
                _pct(q["score_rate"]),
            ])
        flow.append(_table(w_rows, font, col_widths=[35 * mm, 30 * mm, 50 * mm, 25 * mm], header=True))
    else:
        flow.append(Paragraph("无薄弱题（所有题得分率 ≥60%）", styles["CN"]))

    doc.build(flow)
    return buf.getvalue()


def render_student_subject_report_xlsx(report: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    center = Alignment(horizontal="center")

    s = report["student"]
    t = report["totals"]

    # Sheet 1: 个人成绩
    ws = wb.active
    ws.title = "个人成绩"
    rows = [
        ["姓名", s.get("name", "")],
        ["学号", s.get("student_number", "")],
        ["班级", s.get("class_name") or ""],
        ["考试", report["exam"]["name"]],
        ["科目", report["subject"]["name"]],
        ["总分", t["score"]],
        ["满分", t["max_score"]],
        ["得分率", _pct_str(t["score_rate"])],
        ["班级均分", t.get("class_avg")],
        ["班级人数", t.get("class_size", 0)],
    ]
    for r in rows:
        ws.append(r)
    for cell in ws["A"]:
        cell.font = bold
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 30

    # Sheet 2: 各题得失分
    ws2 = wb.create_sheet("各题得失分")
    ws2.append(["题号", "题型", "得分", "满分", "得分率"])
    for cell in ws2[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
    for q in report["questions"]:
        ws2.append([
            q["question_name"], q["question_type"],
            q["score"], q["max_score"], _pct_str(q["score_rate"]),
        ])
    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 14

    # Sheet 3: 薄弱题
    ws3 = wb.create_sheet("薄弱题")
    ws3.append(["题号", "题型", "得分", "满分", "得分率"])
    for cell in ws3[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
    for q in report["weakness"]:
        ws3.append([
            q["question_name"], q["question_type"],
            q["score"], q["max_score"], _pct_str(q["score_rate"]),
        ])
    for col in range(1, 6):
        ws3.column_dimensions[get_column_letter(col)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def build_grade_subject_report(
    db: AsyncSession,
    *,
    exam_id: str,
    subject_id: str,
    school_id: str,
    visible_subject_codes: list[str] | None,
    visible_class_ids: list[str] | None,
) -> dict:
    """聚合年级学科报告所需的全部数据。"""
    exam = (await db.execute(
        select(Exam).where(Exam.id == exam_id, Exam.school_id == school_id)
    )).scalar_one_or_none()
    if not exam:
        raise NotFoundError("Exam not found")

    subject = (await db.execute(
        select(Subject).where(
            Subject.id == subject_id,
            Subject.exam_id == exam_id,
            Subject.school_id == school_id,
        )
    )).scalar_one_or_none()
    if not subject:
        raise NotFoundError("Subject not found")

    if visible_subject_codes is not None and subject.code not in visible_subject_codes:
        from edu_cloud.services.exceptions import PermissionDeniedError
        raise PermissionDeniedError("无权访问该科目")

    summary = await analytics_service.exam_summary(
        db, exam_id=exam_id, school_id=school_id,
        visible_subject_codes=visible_subject_codes,
        visible_class_ids=visible_class_ids,
    )
    subject_summary = next(
        (s for s in summary["subjects"] if s["subject_id"] == subject_id), None,
    )

    distribution = await analytics_service.exam_distribution(
        db, exam_id=exam_id, school_id=school_id, subject_id=subject_id,
        visible_subject_codes=visible_subject_codes,
        visible_class_ids=visible_class_ids,
    )

    aggregates = await analytics_service.grade_aggregates(
        db, exam_id=exam_id, school_id=school_id, subject_id=subject_id,
        visible_subject_codes=visible_subject_codes,
        visible_class_ids=visible_class_ids,
    )

    questions = await analytics_service.subject_question_analysis(
        db, subject_id=subject_id, school_id=school_id,
        visible_subject_codes=visible_subject_codes,
        visible_class_ids=visible_class_ids,
    )

    sorted_qs = sorted(questions["questions"], key=lambda q: q["score_rate"], reverse=True)
    top = sorted_qs[:5]
    bottom = list(reversed(sorted_qs[-5:])) if len(sorted_qs) >= 1 else []

    return {
        "exam": {"id": exam.id, "name": exam.name},
        "subject": {"id": subject.id, "name": subject.name, "code": subject.code},
        "subject_summary": subject_summary,
        "distribution": distribution,
        "class_rankings": aggregates["class_rankings"],
        "grade_stats": aggregates["grade_stats"],
        "questions": questions["questions"],
        "questions_top": top,
        "questions_bottom": bottom,
    }


def render_grade_subject_report_pdf(report: dict) -> bytes:
    """reportlab Platypus 渲染 PDF。"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )

    font = _ensure_cjk_font()
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="CN", parent=styles["Normal"], fontName=font, fontSize=10, leading=14,
    ))
    styles.add(ParagraphStyle(
        name="CN-H1", parent=styles["Heading1"], fontName=font, fontSize=18, leading=22,
        spaceAfter=8,
    ))
    styles.add(ParagraphStyle(
        name="CN-H2", parent=styles["Heading2"], fontName=font, fontSize=13, leading=18,
        spaceBefore=10, spaceAfter=6,
    ))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=20 * mm, rightMargin=20 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
    )
    flow = []
    title = f"{report['exam']['name']} · {report['subject']['name']} 年级分析报告"
    flow.append(Paragraph(title, styles["CN-H1"]))

    # 1. 总览
    flow.append(Paragraph("一、总览", styles["CN-H2"]))
    s = report["subject_summary"] or {}
    summary_rows = [
        ["科目", report["subject"]["name"]],
        ["参加人数", str(s.get("graded_count", 0))],
        ["满分", _fmt(s.get("max_score_possible"))],
        ["平均分", _fmt(s.get("avg_score"))],
        ["最高分", _fmt(s.get("highest"))],
        ["最低分", _fmt(s.get("lowest"))],
        ["得分率", _pct(s.get("score_rate"))],
    ]
    flow.append(_table(summary_rows, font, col_widths=[40 * mm, 60 * mm]))

    # 2. 分数段分布
    flow.append(Paragraph("二、分数段分布", styles["CN-H2"]))
    intervals = report["distribution"].get("intervals", [])
    dist_rows = [["分数段", "区间", "人数"]]
    for itv in intervals:
        dist_rows.append([
            itv.get("label", ""),
            f"{_fmt(itv.get('min'))} ~ {_fmt(itv.get('max'))}",
            str(itv.get("count", 0)),
        ])
    flow.append(_table(dist_rows, font, col_widths=[35 * mm, 50 * mm, 25 * mm], header=True))

    # 3. 班级排名
    flow.append(Paragraph("三、班级对比", styles["CN-H2"]))
    rank_rows = [["排名", "班级", "平均分", "人数"]]
    for c in report["class_rankings"]:
        rank_rows.append([
            str(c.get("rank", "")),
            c.get("class_name", ""),
            _fmt(c.get("avg_score")),
            str(c.get("student_count", 0)),
        ])
    flow.append(_table(rank_rows, font, col_widths=[20 * mm, 50 * mm, 30 * mm, 30 * mm], header=True))

    # 4. 题目分析（top + bottom）
    flow.append(Paragraph("四、题目得分率（高 5 / 低 5）", styles["CN-H2"]))
    q_rows = [["题号", "题型", "满分", "平均分", "得分率"]]
    for q in report["questions_top"] + report["questions_bottom"]:
        q_rows.append([
            q.get("question_name", ""),
            q.get("question_type", ""),
            _fmt(q.get("max_score")),
            _fmt(q.get("avg_score")),
            _pct(q.get("score_rate")),
        ])
    flow.append(_table(q_rows, font, col_widths=[35 * mm, 30 * mm, 25 * mm, 30 * mm, 25 * mm], header=True))

    doc.build(flow)
    return buf.getvalue()


def render_grade_subject_report_xlsx(report: dict) -> bytes:
    """openpyxl 多 sheet 写法（参考 conduct/export_service.py）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    center = Alignment(horizontal="center")

    # Sheet 1: 总览
    ws = wb.active
    ws.title = "总览"
    s = report["subject_summary"] or {}
    rows = [
        ["考试", report["exam"]["name"]],
        ["科目", report["subject"]["name"]],
        ["参加人数", s.get("graded_count", 0)],
        ["满分", s.get("max_score_possible")],
        ["平均分", s.get("avg_score")],
        ["最高分", s.get("highest")],
        ["最低分", s.get("lowest")],
        ["得分率", _pct_str(s.get("score_rate"))],
    ]
    for r in rows:
        ws.append(r)
    for cell in ws["A"]:
        cell.font = bold
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 30

    # Sheet 2: 分数段分布
    ws2 = wb.create_sheet("分数段分布")
    ws2.append(["分数段", "区间下限", "区间上限", "人数"])
    for cell in ws2[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
    for itv in report["distribution"].get("intervals", []):
        ws2.append([
            itv.get("label", ""),
            itv.get("min"),
            itv.get("max"),
            itv.get("count", 0),
        ])
    for col in range(1, 5):
        ws2.column_dimensions[get_column_letter(col)].width = 14

    # Sheet 3: 班级对比
    ws3 = wb.create_sheet("班级对比")
    ws3.append(["排名", "班级", "平均分", "人数", "本班"])
    for cell in ws3[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
    for c in report["class_rankings"]:
        ws3.append([
            c.get("rank", ""),
            c.get("class_name", ""),
            c.get("avg_score"),
            c.get("student_count", 0),
            "★" if c.get("my_class") else "",
        ])
    for col in range(1, 6):
        ws3.column_dimensions[get_column_letter(col)].width = 12

    # Sheet 4: 题目分析
    ws4 = wb.create_sheet("题目分析")
    ws4.append(["题号", "题型", "满分", "平均分", "得分率", "已批改"])
    for cell in ws4[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
    for q in report["questions"]:
        ws4.append([
            q.get("question_name", ""),
            q.get("question_type", ""),
            q.get("max_score"),
            q.get("avg_score"),
            _pct_str(q.get("score_rate")),
            q.get("graded_count", 0),
        ])
    for col in range(1, 7):
        ws4.column_dimensions[get_column_letter(col)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Helpers ─────────────────────────────────────────────────────

def _fmt(v) -> str:
    if v is None:
        return "—"
    if isinstance(v, float):
        return f"{v:.2f}".rstrip("0").rstrip(".") or "0"
    return str(v)


def _pct(v) -> str:
    if v is None:
        return "—"
    return f"{v * 100:.1f}%"


def _pct_str(v) -> str:
    if v is None:
        return ""
    return f"{v * 100:.1f}%"


def _table(rows, font, col_widths, header: bool = False):
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    t = Table(rows, colWidths=col_widths)
    style = [
        ("FONT", (0, 0), (-1, -1), font, 9),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
    ]
    if header:
        style.append(("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey))
        style.append(("FONT", (0, 0), (-1, 0), font, 10))
    t.setStyle(TableStyle(style))
    return t
