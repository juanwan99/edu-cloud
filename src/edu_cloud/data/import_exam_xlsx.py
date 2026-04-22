"""导入真实考试 Excel 数据到 edu-cloud。

数据源：株洲市二中枫溪学校 2026 届高三一模（3月4-5号考试）
- 总分(物理)_学生成绩(方向名次).xlsx — 物理方向 257 人
- 总分(历史)_学生成绩(方向名次).xlsx — 历史方向 110 人

导入内容：
1. 学校（如不存在则创建）
2. 班级（2301-2309）
3. 学生（367 人）
4. 考试 + 科目
5. 每个学生每科的成绩（ExamResult + 分科成绩）

调用：python -m edu_cloud.data.import_exam_xlsx
幂等：检查考试名是否已存在。
"""
import asyncio
import logging
import os
import sys

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker, AsyncSession

logger = logging.getLogger(__name__)

DATA_DIR = "D:/试卷数据/试卷图像/191871"
SCHOOL_NAME = "株洲市二中枫溪学校"
SCHOOL_CODE = "ZZSEZ2026"
EXAM_NAME = "2026届高三3月一模"

# 科目满分
SUBJECT_MAX_SCORES = {
    "语文": 150, "数学": 150, "英语": 150,
    "物理": 100, "化学": 100, "生物": 100,
    "政治": 100, "历史": 100, "地理": 100,
}

# 科目在 Excel 中的列位置（原始分列）
# 两个文件结构相同：col13=语文, col17=数学, col21=英语, col25=方向主科(物理/历史)
SCORE_COLUMNS = {
    "语文": 13, "数学": 17, "英语": 21,
}

# 赋分科目的原始分列（在赋分列后面 5 列）
# 物理方向文件：col25=物理, col29=化学赋分→col34=化学原始, col38=生物赋分→col43=生物原始
#               col47=政治赋分→col52=政治原始, col56=地理赋分→col61=地理原始
# 历史方向文件：col25=历史, 后面同理

PHYSICS_FILE_SCORES = {
    "物理": 25,
    "化学": 34, "生物": 43,
    "政治": 52, "地理": 61,
}

HISTORY_FILE_SCORES = {
    "历史": 25,
    "化学": 34, "生物": 43,
    "政治": 52, "地理": 61,
}


def _read_students(filepath, direction_scores):
    """读取一个 Excel 文件，返回学生数据列表。"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    students = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        if not row[1] or not row[3]:  # 无班级或无姓名则跳过
            continue

        name = str(row[3]).strip()
        class_name = str(row[1]).strip()
        exam_number = str(row[2]).strip() if row[2] else ""
        subject_combo = str(row[4]).strip() if row[4] else ""

        # 总分（赋分）
        total_score = float(row[5]) if row[5] else 0
        total_raw = float(row[9]) if row[9] else 0

        # 各科成绩
        scores = {}
        # 共同科目（语数英）
        for subj, col in SCORE_COLUMNS.items():
            val = row[col] if col < len(row) and row[col] is not None else None
            if val is not None:
                scores[subj] = float(val)

        # 方向特有科目
        for subj, col in direction_scores.items():
            val = row[col] if col < len(row) and row[col] is not None else None
            if val is not None:
                scores[subj] = float(val)

        students.append({
            "name": name,
            "class_name": class_name,
            "exam_number": exam_number,
            "subject_combo": subject_combo,
            "total_score": total_score,
            "total_raw": total_raw,
            "scores": scores,
        })

    return students


async def import_exam_data(db: AsyncSession) -> dict:
    """导入考试数据。"""
    from edu_cloud.models.school import School
    from edu_cloud.models.user import User
    from edu_cloud.models.user_role import UserRole
    from edu_cloud.modules.student.models import Class, Student
    from edu_cloud.modules.exam.models import Exam, Subject, ExamResult

    # 幂等检查
    existing = (await db.execute(
        select(Exam).where(Exam.name == EXAM_NAME)
    )).scalar_one_or_none()
    if existing:
        return {"status": "already_imported", "message": f"考试 '{EXAM_NAME}' 已存在"}

    # ── 1. 学校 ──
    school = (await db.execute(
        select(School).where(School.code == SCHOOL_CODE)
    )).scalar_one_or_none()
    if not school:
        school = School(name=SCHOOL_NAME, code=SCHOOL_CODE, district="株洲市")
        db.add(school)
        await db.flush()
        logger.info("Created school: %s", SCHOOL_NAME)
    school_id = school.id

    # ── 2. 读取 Excel ──
    physics_file = os.path.join(DATA_DIR, "总分(物理)_学生成绩(方向名次).xlsx")
    history_file = os.path.join(DATA_DIR, "总分(历史)_学生成绩(方向名次).xlsx")

    physics_students = _read_students(physics_file, PHYSICS_FILE_SCORES)
    history_students = _read_students(history_file, HISTORY_FILE_SCORES)
    all_students = physics_students + history_students

    logger.info("Read %d physics + %d history = %d students",
                len(physics_students), len(history_students), len(all_students))

    # ── 3. 创建班级 ──
    class_names = sorted(set(s["class_name"] for s in all_students))
    class_map = {}  # class_name → Class obj
    for cname in class_names:
        cls = (await db.execute(
            select(Class).where(Class.name == cname, Class.school_id == school_id)
        )).scalar_one_or_none()
        if not cls:
            cls = Class(name=cname, grade="高三", grade_number=12, school_id=school_id)
            db.add(cls)
        class_map[cname] = cls
    await db.flush()
    logger.info("Created/found %d classes: %s", len(class_map), class_names)

    # ── 4. 创建学生 ──
    student_map = {}  # exam_number → Student obj
    for s in all_students:
        student = (await db.execute(
            select(Student).where(
                Student.student_number == s["exam_number"],
                Student.school_id == school_id,
            )
        )).scalar_one_or_none()
        if not student:
            student = Student(
                name=s["name"],
                student_number=s["exam_number"],
                class_id=class_map[s["class_name"]].id,
                school_id=school_id,
                grade="高三",
                status="active",
            )
            db.add(student)
        student_map[s["exam_number"]] = student
    await db.flush()
    logger.info("Created/found %d students", len(student_map))

    # ── 5. 创建考试 + 科目 ──
    exam = Exam(
        name=EXAM_NAME,
        card_title="2026届高三一模",
        status="completed",
        exam_type="monthly",
        grade_scope="高三",
        semester="2025-2026-2",
        school_id=school_id,
        max_score=750,  # 3+1+2 模式总分
    )
    db.add(exam)
    await db.flush()

    # 收集所有出现过的科目
    all_subjects = set()
    for s in all_students:
        all_subjects.update(s["scores"].keys())

    subject_map = {}  # subject_name → Subject obj
    for subj_name in sorted(all_subjects):
        code = {
            "语文": "YW", "数学": "SX", "英语": "YY",
            "物理": "WL", "化学": "HX", "生物": "SW",
            "政治": "ZZ", "历史": "LS", "地理": "DL",
        }.get(subj_name, subj_name[:2].upper())

        subject = Subject(
            name=subj_name,
            code=code,
            exam_id=exam.id,
            school_id=school_id,
        )
        db.add(subject)
        subject_map[subj_name] = subject
    await db.flush()
    logger.info("Created %d subjects: %s", len(subject_map), list(subject_map.keys()))

    # ── 6. 写入成绩 ──
    result_count = 0
    for s in all_students:
        student = student_map[s["exam_number"]]

        # 总分 ExamResult
        exam_result = ExamResult(
            exam_id=exam.id,
            student_id=student.id,
            school_id=school_id,
            total_score=s["total_raw"],
        )
        db.add(exam_result)
        result_count += 1

    await db.commit()
    logger.info("Created %d exam results", result_count)

    return {
        "status": "imported",
        "school": SCHOOL_NAME,
        "exam": EXAM_NAME,
        "classes": len(class_map),
        "students": len(student_map),
        "subjects": len(subject_map),
        "results": result_count,
    }


# ── 5b. 创建教师用户（给这所学校的教师角色） ──

async def create_school_users(db: AsyncSession, school_id: str):
    """为导入的学校创建校长和几名教师用户，方便登录测试。"""
    from edu_cloud.models.user import User
    from edu_cloud.models.user_role import UserRole

    users_to_create = [
        ("zhao_principal", "赵校长", "principal"),
        ("wang_director", "王教务", "academic_director"),
        ("li_teacher_yw", "李语文", "subject_teacher"),
        ("zhang_teacher_sx", "张数学", "homeroom_teacher"),
    ]

    created = 0
    for username, display_name, role in users_to_create:
        existing = (await db.execute(
            select(User).where(User.username == username)
        )).scalar_one_or_none()
        if existing:
            continue

        user = User(username=username, display_name=display_name)
        user.set_password("123456")
        db.add(user)
        await db.flush()

        ur = UserRole(
            user_id=user.id,
            role=role,
            school_id=school_id,
            is_primary=True,
        )
        if role == "subject_teacher":
            ur.subject_codes = ["YW"]
        elif role == "homeroom_teacher":
            ur.subject_codes = ["SX"]
            ur.class_ids = []  # 将在后面填充
        db.add(ur)
        created += 1

    await db.commit()
    return created


async def main():
    src_dir = os.path.join(os.path.dirname(__file__), "..", "..")
    sys.path.insert(0, os.path.abspath(src_dir))

    from edu_cloud.config import settings
    engine = create_async_engine(settings.DATABASE_URL, echo=False)

    if "sqlite" in settings.DATABASE_URL:
        from edu_cloud.models.base import Base
        import edu_cloud.models.school  # noqa
        import edu_cloud.models.user  # noqa
        import edu_cloud.models.user_role  # noqa
        import edu_cloud.models.student  # noqa
        import edu_cloud.models.class_group  # noqa
        import edu_cloud.models.exam  # noqa
        import edu_cloud.models.ai_session  # noqa
        import edu_cloud.models.document  # noqa
        import edu_cloud.models.approval  # noqa
        import edu_cloud.models.calendar  # noqa
        import edu_cloud.models.notification  # noqa
        import edu_cloud.models.llm_slot  # noqa
        import edu_cloud.modules.card.models  # noqa
        import edu_cloud.modules.scan.models  # noqa
        import edu_cloud.modules.grading.models  # noqa
        import edu_cloud.modules.marking.models  # noqa
        import edu_cloud.modules.knowledge.models  # noqa
        import edu_cloud.modules.bank.models  # noqa
        import edu_cloud.modules.profile.models  # noqa

        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

    session_factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with session_factory() as session:
        result = await import_exam_data(session)

        print(f"\n{'=' * 50}")
        for k, v in result.items():
            print(f"  {k}: {v}")
        print(f"{'=' * 50}")

        if result["status"] == "imported":
            # 创建测试用户
            from sqlalchemy import select as sel
            from edu_cloud.models.school import School
            school = (await session.execute(
                sel(School).where(School.code == SCHOOL_CODE)
            )).scalar_one()
            n = await create_school_users(session, school.id)
            print(f"  Created {n} test users for {SCHOOL_NAME}")

    await engine.dispose()


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    asyncio.run(main())
