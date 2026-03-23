"""导入真实考试数据 — 从 Excel 读取学生信息、班级、小题分数。

数据来源：D:/试卷数据/生物.xlsx
格式：准考证号 | 自定义考号 | 班级 | 姓名 | 总分 | 等级 | 校次 | 班次 | 客观分 | 主观分 | 单选题 | 多选题 | 解答题 | 第1题 | ... | 第21题

导入内容：
1. School（如果不存在）
2. Classes（从班级列去重）
3. Students（从学生列去重，按准考证号匹配）
4. Exam + Subject + Questions（根据题目数和分值创建）
5. StudentAnswer（每个学生每道题的分数）

Migrated from exam-ai (Task 22).
"""
import logging
import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from edu_cloud.models.school import School
from edu_cloud.modules.student.models import Class, Student
from edu_cloud.modules.exam.models import Exam, Subject, Question
from edu_cloud.modules.scan.models import StudentAnswer

logger = logging.getLogger(__name__)

# 生物题目满分推断：第1-16题选择题每题2分，第17-21题主观题
# 主观题满分需从数据推断
OBJECTIVE_MAX = 2  # 选择题每题2分
OBJECTIVE_COUNT = 16
SUBJECTIVE_START = 17  # 第17题开始是主观题
TOTAL_QUESTIONS = 21


async def import_biology_exam(
    db: AsyncSession,
    excel_path: str = "",  # 由调用方传入，支持 Windows 或 WSL 路径
    school_name: str = "测试学校",
    school_code: str = "TEST01",  # 导入到现有 seed 学校，admin 可见
    exam_name: str = "2026届高三3月模考",
    force_reimport: bool = False,
) -> dict:
    """导入生物考试数据。返回统计信息。force_reimport=True 删除旧数据重新导入。"""

    # === 读 Excel ===
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]

    # 解析数据行（从第3行开始，跳过汇总行）
    rows = []
    for row_idx in range(3, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=4).value
        if not name or str(name).strip() in ("-", ""):
            continue
        # 跳过"缺考"行和汇总行（total_score 不是数字）
        total_raw = ws.cell(row=row_idx, column=5).value
        if total_raw and isinstance(total_raw, str) and "缺考" in total_raw:
            continue
        class_name = ws.cell(row=row_idx, column=3).value
        if not class_name or str(class_name).strip() == "":
            continue  # 无班级 = 汇总行
        row = {
            "exam_ticket": str(ws.cell(row=row_idx, column=1).value or ""),
            "custom_id": str(ws.cell(row=row_idx, column=2).value or ""),
            "class_name": str(ws.cell(row=row_idx, column=3).value or "").strip(),
            "name": str(name).strip(),
            "total_score": _safe_float(ws.cell(row=row_idx, column=5).value),
            "school_rank": _safe_int(ws.cell(row=row_idx, column=7).value),
            "class_rank": _safe_int(ws.cell(row=row_idx, column=8).value),
            "objective_score": _safe_float(ws.cell(row=row_idx, column=9).value),
            "subjective_score": _safe_float(ws.cell(row=row_idx, column=10).value),
            "question_scores": [],
        }
        for q_col in range(14, 14 + TOTAL_QUESTIONS):  # 1-indexed col 14 = 第1题
            val = ws.cell(row=row_idx, column=q_col).value
            row["question_scores"].append(_safe_float(val))
        rows.append(row)

    if not rows:
        return {"status": "error", "message": "Excel 中没有有效数据行"}

    # === 推断主观题满分 ===
    subjective_max = {}
    for q_idx in range(SUBJECTIVE_START - 1, TOTAL_QUESTIONS):
        max_val = max(r["question_scores"][q_idx] for r in rows)
        subjective_max[q_idx + 1] = max_val  # 题号从1开始

    # === 创建/获取 School ===
    result = await db.execute(select(School).where(School.code == school_code))
    school = result.scalar_one_or_none()
    if not school:
        school = School(name=school_name, code=school_code)
        db.add(school)
        await db.flush()
        logger.info("created school: %s (%s)", school_name, school_code)

    # === 创建/获取 Classes ===
    class_names = sorted(set(r["class_name"] for r in rows if r["class_name"]))
    class_map = {}  # name → Class
    for cname in class_names:
        result = await db.execute(
            select(Class).where(Class.name == cname, Class.school_id == school.id)
        )
        cls = result.scalar_one_or_none()
        if not cls:
            # 从班级名提取年级信息
            grade = "高三" if "高三" in cname else "高二" if "高二" in cname else "未知"
            cls = Class(name=cname, grade=grade, school_id=school.id)
            db.add(cls)
            await db.flush()
        class_map[cname] = cls
    logger.info("classes: %d total, names=%s", len(class_map), list(class_map.keys()))

    # === 创建/获取 Students ===
    student_map = {}  # custom_id → Student
    created_students = 0
    for r in rows:
        student_number = r["custom_id"] or r["exam_ticket"]
        if not student_number or student_number.strip() in ("-", ""):
            continue
        result = await db.execute(
            select(Student).where(
                Student.student_number == student_number,
                Student.school_id == school.id,
            )
        )
        student = result.scalar_one_or_none()
        if not student:
            cls = class_map.get(r["class_name"])
            student = Student(
                name=r["name"],
                student_number=student_number,
                class_id=cls.id if cls else list(class_map.values())[0].id,
                school_id=school.id,
            )
            db.add(student)
            created_students += 1
        student_map[student_number] = student
    await db.flush()
    logger.info("students: %d total, %d new", len(student_map), created_students)

    # === 创建 Exam + Subject ===
    result = await db.execute(
        select(Exam).where(Exam.name == exam_name, Exam.school_id == school.id)
    )
    exam = result.scalar_one_or_none()
    if exam and not force_reimport:
        return {
            "status": "already_exists",
            "message": f"考试 '{exam_name}' 已存在 (id={exam.id})，传 force_reimport=True 重新导入",
            "school": school_name,
            "students": len(student_map),
        }
    if exam and force_reimport:
        # 删除旧考试的关联数据
        from edu_cloud.modules.grading.models import AIGradingResult, GradingTask
        from edu_cloud.modules.bank.models import StudentErrorBook, BankQuestion
        from edu_cloud.modules.profile.models import StudentExamSnapshot, StudentErrorPattern
        logger.info("force_reimport: deleting old exam data for '%s'", exam_name)
        await db.execute(AIGradingResult.__table__.delete().where(AIGradingResult.__table__.c.school_id == school.id))
        await db.execute(GradingTask.__table__.delete().where(GradingTask.__table__.c.school_id == school.id))
        await db.execute(StudentAnswer.__table__.delete().where(StudentAnswer.__table__.c.exam_id == exam.id))
        await db.execute(StudentErrorBook.__table__.delete().where(StudentErrorBook.__table__.c.exam_id == exam.id))
        await db.execute(BankQuestion.__table__.delete().where(BankQuestion.__table__.c.source_exam_id == exam.id))
        await db.execute(StudentExamSnapshot.__table__.delete().where(StudentExamSnapshot.__table__.c.exam_id == exam.id))
        await db.execute(StudentErrorPattern.__table__.delete().where(StudentErrorPattern.__table__.c.school_id == school.id))
        await db.execute(Subject.__table__.delete().where(Subject.__table__.c.exam_id == exam.id))
        await db.execute(Exam.__table__.delete().where(Exam.__table__.c.id == exam.id))
        await db.commit()
        logger.info("force_reimport: old data deleted")

    exam = Exam(
        name=exam_name, card_title=exam_name,
        school_id=school.id, status="completed",
        exam_type="模拟", grade_scope="高三",
        semester="2025-2026-2",
    )
    db.add(exam)
    await db.flush()

    subject = Subject(
        exam_id=exam.id, name="生物", code="SW",
        school_id=school.id,
    )
    db.add(subject)
    await db.flush()

    # 创建 GradingTask（AIGradingResult 需要 task_id FK）
    from edu_cloud.modules.grading.models import GradingTask
    from edu_cloud.models.user import User
    from edu_cloud.models.user_role import UserRole
    # edu-cloud User 没有 school_id/role，通过 UserRole 查找 admin
    admin_role_result = await db.execute(
        select(UserRole).where(
            UserRole.school_id == school.id,
            UserRole.role == "platform_admin",
        )
    )
    admin_role = admin_role_result.scalar_one_or_none()
    admin_user_id = admin_role.user_id if admin_role else school.id
    grading_task = GradingTask(
        subject_id=subject.id, school_id=school.id,
        status="completed", total=0, completed=0, failed=0,
        created_by=admin_user_id,
    )
    db.add(grading_task)
    await db.flush()

    # === 创建 Questions ===
    questions = []
    for q_num in range(1, TOTAL_QUESTIONS + 1):
        if q_num <= OBJECTIVE_COUNT:
            q_type = "objective"
            max_score = float(OBJECTIVE_MAX)
        else:
            q_type = "subjective"
            max_score = float(subjective_max.get(q_num, 12))
        q = Question(
            subject_id=subject.id, school_id=school.id,
            name=str(q_num), question_type=q_type,
            max_score=max_score,
        )
        db.add(q)
        questions.append(q)
    await db.flush()
    logger.info("questions: %d created (obj=%d, subj=%d)",
                len(questions), OBJECTIVE_COUNT, TOTAL_QUESTIONS - OBJECTIVE_COUNT)

    # === 创建 StudentAnswer ===
    answer_count = 0
    seen_keys = set()
    for r in rows:
        student_number = r["custom_id"] or r["exam_ticket"]
        if not student_number or student_number.strip() in ("-", ""):
            continue
        student = student_map.get(student_number)
        if not student:
            continue

        for q_idx, q in enumerate(questions):
            # 去重：同一学生同一题只插一次
            dedup_key = (student.id, q.id)
            if dedup_key in seen_keys:
                continue
            seen_keys.add(dedup_key)

            score = r["question_scores"][q_idx]
            detected = None
            if q.question_type == "objective":
                detected = "correct" if score >= q.max_score else "wrong"

            sa = StudentAnswer(
                exam_id=exam.id, subject_id=subject.id,
                student_id=student.id, question_id=q.id,
                school_id=school.id, score=score,
                detected_answer=detected,
            )
            db.add(sa)
            await db.flush()  # 需要 sa.id 给 AIGradingResult

            # 创建 AIGradingResult（analytics 的 get_effective_scores 通过 JOIN AIGradingResult 获取成绩）
            from edu_cloud.modules.grading.models import AIGradingResult
            ai_result = AIGradingResult(
                task_id=grading_task.id,
                answer_id=sa.id,
                question_id=q.id,
                school_id=school.id,
                score=score,
                max_score=q.max_score,
                feedback=f"Excel 导入成绩: {score}/{q.max_score}",
                confidence=1.0,
                review_status="approved",
            )
            db.add(ai_result)
            answer_count += 1

    await db.commit()
    logger.info("import complete: exam=%s, answers=%d", exam_name, answer_count)

    # === 运行 data pipeline ===
    pipeline_results = None
    try:
        from edu_cloud.modules.pipeline.service import run_full_pipeline
        pipeline_results = await run_full_pipeline(db, exam_id=exam.id, school_id=school.id)
        logger.info("pipeline results: %s", pipeline_results)
    except Exception as e:
        logger.error("pipeline failed: %s", e, exc_info=True)
        pipeline_results = {"error": str(e)}

    return {
        "status": "success",
        "school": school_name,
        "exam": exam_name,
        "classes": len(class_map),
        "students_total": len(student_map),
        "students_new": created_students,
        "questions": len(questions),
        "answers": answer_count,
        "pipeline": pipeline_results,
    }


def _safe_float(val) -> float:
    if val is None or str(val).strip() in ("-", "", "None", "缺考"):
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _safe_int(val) -> int | None:
    if val is None or str(val).strip() in ("-", "", "None"):
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None
